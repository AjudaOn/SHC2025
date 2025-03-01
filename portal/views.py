from datetime import datetime
from io import BytesIO
from unittest import case
from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django import forms
from django.core.paginator import Paginator
from django.views import View
from portal.models import BaseDados, Produto, Precos_graduacao_vinculo, Precos_status_graduacao
from .forms import ReservasForm
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseNotAllowed, JsonResponse
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect
import logging
from decimal import Decimal
from django.template.defaulttags import register
from django.db.models import Sum
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import locale
from datetime import datetime
from django.db.models import F
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
import xlsxwriter
from django.http import Http404
from django.db.models import Case, When, Value, F, IntegerField, Q
from django.views.generic import View
import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from datetime import timedelta
from django.shortcuts import render
from django.utils import timezone
from calendar import monthrange
from django.db.models import Avg
from datetime import datetime, timedelta, date
import decimal
from django.db import transaction
from django.http import JsonResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt



@register.filter
def get_range(value):
    return range(value)


@login_required
def consultar_reservas(request):
    consultar = BaseDados.objects.filter(status_reserva="Pendente")  # Filtra por status "pendente"
    print(consultar)
    context = {
        'consultar': consultar,
    }
    return render(request, 'portal/reservas.html', context)

# def reserva_externa(request):
#     if request.method == 'POST':
#         form = ReservasForm(request.POST)
#         if form.is_valid():
#             obj = form.save()
#             print("Objeto salvo:", obj)
#             form.save()
#             # return redirect('consultar_reservas')
#             return redirect('reserva_externa')
#         else:
#             # Se o formulário for inválido, cairemos aqui e os erros serão impressos.
#             print("Formulário inválido", form.errors)
#     else:
#         # Se a requisição não for POST, um formulário em branco será fornecido.
#         form = ReservasForm()

#     # O contexto é passado para o template tanto se o formulário for válido (e redirecionado antes),
#     # quanto se não for POST ou se for inválido.
#     context = {'form': form}
#     return render(request, 'portal/index.html', context)


# @transaction.atomic
# def reserva_externa(request):
#     if request.method == 'POST':
#         form = ReservasForm(request.POST)
#         if form.is_valid():
#             obj = form.save()
#             print("Objeto salvo:", obj, "ID:", obj.id)
#             return redirect('reserva_externa')
#         else:
#             print("Formulário inválido", form.errors)
#     else:
#         form = ReservasForm()

#     context = {'form': form}
#     return render(request, 'portal/index.html', context)

@transaction.atomic
def reserva_externa(request):
    if request.method == 'POST':
        form = ReservasForm(request.POST)
        if form.is_valid():
            try:
                obj = form.save()  # Salva o formulário e obtém o objeto salvo
                print("Objeto salvo:", obj, "ID:", obj.id)
                return JsonResponse({"success": True, "message": "Cadastro realizado com sucesso!"})
            except Exception as e:
                # Em caso de erro no salvamento, retorna uma resposta de falha
                print("Erro ao salvar:", str(e))
                return JsonResponse({"success": False, "message": "Erro ao salvar os dados!"}, status=500)
        else:
            # Se o formulário for inválido, retorna os erros em formato JSON
            errors = form.errors.as_json()
            print("Formulário inválido", errors)
            return JsonResponse({"success": False, "message": "Dados do formulário inválidos!", "errors": errors}, status=400)
    else:
        # Se a requisição não for POST, retorna um HTML normal
        form = ReservasForm()
        context = {'form': form}
        return render(request, 'portal/index.html', context)



from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import ReservasForm

from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .forms import ReservasForm

from django.utils.timezone import now
from django.shortcuts import render, redirect, get_object_or_404
from .forms import ReservasForm
from django.contrib.auth.decorators import login_required

@login_required
def fazer_reservas(request):
    if request.method == 'POST':
        form = ReservasForm(request.POST)
        if form.is_valid():
            # Validação adicional para as datas
            entrada = form.cleaned_data.get('entrada')
            saida = form.cleaned_data.get('saida')
            hoje = now().date()

            if entrada and entrada < hoje:
                form.add_error('entrada', 'A data de entrada não pode ser anterior à data de hoje.')
            if saida and saida < hoje:
                form.add_error('saida', 'A data de saída não pode ser anterior à data de hoje.')
            if entrada and saida and entrada > saida:
                form.add_error('saida', 'A data de saída deve ser posterior ou igual à data de entrada.')

            # Se houver erros na validação, retorna o formulário com os erros
            if form.errors:
                return render(request, 'portal/reserva_add.html', {'form': form})

            # Se tudo estiver válido, salva a reserva
            form.save()
            return redirect('consultar_reservas')
    else:
        form = ReservasForm()

    return render(request, 'portal/reserva_add.html', {'form': form})



    
from django.utils.timezone import localdate

# @login_required
# def editar_reservas(request, reservas_pk):
#     editar = get_object_or_404(BaseDados, pk=reservas_pk)

#     if request.method == 'POST':
#         form = ReservasForm(request.POST, instance=editar)

#         # Mantém a lógica de status_reserva
#         if 'status_reserva' in request.POST:
#             status_reserva = request.POST['status_reserva']
#             form.instance.status_reserva = status_reserva

#         if form.is_valid():
#             form.save()
#             return redirect('consultar_reservas')

#     else:
#         form = ReservasForm(instance=editar)

#         # Formatar datas para o formato esperado pelo <input type="date">
#         if editar.entrada:
#             form.fields['entrada'].initial = editar.entrada.strftime('%Y-%m-%d')
#         if editar.saida:
#             form.fields['saida'].initial = editar.saida.strftime('%Y-%m-%d')

#     context = {'form': form, 'reserva': editar, 'modo': 'editar'}
#     return render(request, 'portal/reserva_edit.html', context)



# @login_required
# def editar_reservas(request, reservas_pk):
#     editar = get_object_or_404(BaseDados, pk=reservas_pk)
    
#     if request.method == 'POST':
#         form = ReservasForm(request.POST, instance=editar)
#         if form.is_valid():
#             form.save()
#             return redirect('consultar_reservas')
#     else:
#         form = ReservasForm(instance=editar)

#     # Formatar datas para exibição correta no input date (YYYY-MM-DD)
#     if editar.entrada:
#         form.fields['entrada'].initial = editar.entrada.strftime('%Y-%m-%d')
#     if editar.saida:
#         form.fields['saida'].initial = editar.saida.strftime('%Y-%m-%d')

#     context = {'form': form, 'reserva': editar, 'modo': 'editar'}
#     return render(request, 'portal/reserva_add.html', context)  

# @login_required
# def fazer_reservas(request):
#     if request.method == 'POST':
#         form = ReservasForm(request.POST)
#         if form.is_valid():
#             # Validação adicional para as datas
#             entrada = form.cleaned_data.get('entrada')
#             saida = form.cleaned_data.get('saida')
#             hoje = now().date()

#             if entrada < hoje:
#                 form.add_error('entrada', 'A data de entrada não pode ser anterior à data de hoje.')
#             if saida < hoje:
#                 form.add_error('saida', 'A data de saída não pode ser anterior à data de hoje.')
#             if entrada and saida and entrada > saida:
#                 form.add_error('saida', 'A data de saída deve ser posterior ou igual à data de entrada.')

#             # Se houver erros após as validações adicionais, retorna ao formulário com os dados preenchidos
#             if form.errors:
#                 return render(request, 'portal/reserva_add.html', {'form': form})

#             # Se tudo estiver válido, salva o formulário e redireciona
#             obj = form.save()
#             return redirect('consultar_reservas')
#         else:
#             # Se o formulário for inválido
#             print("Formulário inválido", form.errors)
#     else:
#         # Se a requisição não for POST, retorna o formulário vazio
#         form = ReservasForm()

#     return render(request, 'portal/reserva_add.html', {'form': form})



# @login_required
# def editar_reservas(request, reservas_pk):
#     editar = get_object_or_404(BaseDados, pk=reservas_pk)
    
#     if request.method == 'POST':
#         form = ReservasForm(request.POST, instance=editar)
#         print("Dados recebidos do formulário:", request.POST)  # Adiciona mensagem de log para verificar os dados do formulário
#         if 'status_reserva' in request.POST:
#             status_reserva = request.POST['status_reserva']
#             print("Status da reserva recebido:", status_reserva)  # Adiciona mensagem de log para verificar o status da reserva
#             form.instance.status_reserva = status_reserva
        
#         if form.is_valid():
#             form.save()
#             print("Formulário válido. Salvando alterações...")  # Adiciona mensagem de log para verificar se o formulário é válido
#             return redirect('consultar_reservas')
#         else:
#             print("Formulário inválido. Erros:", form.errors)  # Adiciona mensagem de log para verificar os erros de validação do formulário
#             # Se o formulário for inválido, podemos retornar uma resposta HTTP com os erros para ajudar na depuração
#             return HttpResponse("Formulário inválido. Erros: {}".format(form.errors), status=400)
#     else:
#         form = ReservasForm(instance=editar)
    
#     context = {'form': form, 'reserva': editar}
#     return render(request, 'portal/reserva_edit.html', context)

from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import BaseDados
from .forms import ReservasForm

@login_required
def editar_reservas(request, reservas_pk):
    editar = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        # Criar uma cópia dos dados do request.POST para edição
        post_data = request.POST.copy()

        # Atualiza apenas os campos editáveis, mantendo os outros valores do banco
        if "entrada" in post_data and post_data["entrada"]:
            editar.entrada = post_data["entrada"]
        if "saida" in post_data and post_data["saida"]:
            editar.saida = post_data["saida"]
        if "status_reserva" in post_data and post_data["status_reserva"]:
            editar.status_reserva = post_data["status_reserva"]
        if "mhex" in post_data and post_data["mhex"]:
            editar.mhex = post_data["mhex"]
        if "uh" in post_data and post_data["uh"]:
            editar.uh = post_data["uh"]

        # Salva apenas os campos alterados
        editar.save()
        
        return redirect('consultar_reservas')

    # Criar um formulário com os dados existentes para exibição
    form = ReservasForm(instance=editar)

    context = {'form': form, 'reserva': editar}
    return render(request, 'portal/reserva_edit.html', context)


@login_required
def recepcao_checkin(request):
    consultar = BaseDados.objects.filter(status_reserva="Aprovada")  # Filtra por status "pendente"
    consultar_checkin = BaseDados.objects.filter(status_reserva="Checkin")
    # print(consultar)    
    context = {
        'consultar': consultar,
        'consultar_checkin': consultar_checkin,
    }
    return render(request, 'portal/recepcao_checkin.html', context)  

@login_required
def recepcao_checkout(request):
    consultar = BaseDados.objects.filter(status_reserva="Aprovada")  # Filtra por status "pendente"
    consultar_checkin = BaseDados.objects.filter(status_reserva="Checkin")
    # print(consultar)    
    context = {
        'consultar': consultar,
        'consultar_checkin': consultar_checkin,
    }
    return render(request, 'portal/recepcao_checkout.html', context)      



@csrf_protect
@login_required
def editar_checkin(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        novo_status = request.POST.get('status_reserva')
        reserva.status_reserva = novo_status
        
        # Calcula as diárias
        if reserva.entrada and reserva.saida:
            try:
                formato_data = "%d/%m/%Y"
                data_entrada_str = reserva.entrada.strftime(formato_data)  # Convertendo para string
                data_saida_str = reserva.saida.strftime(formato_data)  # Convertendo para string

                data_entrada = datetime.datetime.strptime(data_entrada_str, formato_data)
                data_saida = datetime.datetime.strptime(data_saida_str, formato_data)
                
                diarias = (data_saida - data_entrada).days
                reserva.diarias = diarias
                print(data_entrada)
                print(data_saida)
                print(reserva.diarias)
            except ValueError as e:
                print(f"Erro ao calcular diárias: {e}")
                reserva.diarias = 0
        
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao_checkin'))
    
    return render(request, 'portal/checkin.html', {'reserva': reserva})



def editar_checkout(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        novo_status = request.POST.get('status_reserva')
        novo_formapg = request.POST.get('forma_pagamento')
        reserva.status_reserva = novo_status
        reserva.forma_pagamento = novo_formapg
        
        # Verifica se o checkbox foi marcado
        if 'pagante_checkbox' in request.POST:
            print("Checkbox foi marcado")
            nome_pagante = request.POST.get('nome_pagante')
            cpf_pagante = request.POST.get('cpf_pagante')
            print("Nome do pagante:", nome_pagante)
            print("CPF do pagante:", cpf_pagante)
        else:
            # Se o checkbox não foi marcado, usa os dados existentes
            print("Checkbox não foi marcado")
            nome_pagante = reserva.nome
            cpf_pagante = reserva.cpf
        
        # Atualiza os campos na reserva
        reserva.nome_pagante = nome_pagante
        reserva.cpf_pagante = cpf_pagante
        
        # Salva a reserva
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao_checkout'))
    
    return render(request, 'portal/checkout.html', {'reserva': reserva})



@csrf_protect

def editar_consumacao(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)

    print("Valores do objeto reserva antes de renderizar o formulário:", reserva.qtde_agua, reserva.qtde_refri, reserva.qtde_cerveja, reserva.status, reserva.graduacao, reserva.valor_hosp)
    
    if request.method == 'POST':
        novo_agua = request.POST.get('qtde_agua')
        novo_refri = request.POST.get('qtde_refri')
        novo_cerveja = request.POST.get('qtde_cerveja')
        novo_pet = request.POST.get('qtde_pet')

        status = reserva.status
        graduacao = reserva.graduacao

        # Consulta os preços com base no status e na graduação
        try:
            if status == 'CIVIL':
                preco_status_graduacao = Precos_status_graduacao.objects.get(status=status, graduacao=graduacao)
                reserva.valor_hosp = preco_status_graduacao.valor
            elif status in ['MILITAR DA ATIVA', 'MILITAR DA RESERVA', 'DEP ACOMPANHADO', 'DEP DESACOMPANHADO', 'PENSIONISTA']:                
                preco_status_graduacao = Precos_status_graduacao.objects.get(status=status, graduacao=graduacao)
                reserva.valor_hosp = preco_status_graduacao.valor
            else:
                # Se não encontrar um preço correspondente, define valor_hosp como zero
                reserva.valor_hosp = 0  # Ou algum outro valor padrão que faça sentido em seu contexto
        except Precos_status_graduacao.DoesNotExist:
            # Se o objeto não for encontrado, define valor_hosp como zero
            reserva.valor_hosp = 0
        
        # Exemplo de cálculo para um acompanhante
        vinculo_acomps = [reserva.vinculo_acomp1, reserva.vinculo_acomp2, reserva.vinculo_acomp3, reserva.vinculo_acomp4, reserva.vinculo_acomp5]
        for i, vinculo_acomp in enumerate(vinculo_acomps):
            if vinculo_acomp:
                try:
                    # Consulta o preço do vínculo e graduação na tabela Precos_graduacao_vinculo
                    preco_vinculo = Precos_graduacao_vinculo.objects.get(graduacao=graduacao, vinculo=vinculo_acomp)
                    setattr(reserva, f'valor_acomp{i+1}', preco_vinculo.valor)
                except Precos_graduacao_vinculo.DoesNotExist:
                    setattr(reserva, f'valor_acomp{i+1}', 0)
        # Cálculo dos novos valores conforme solicitado
        if isinstance(reserva.valor_hosp, (int, float, Decimal)):
            reserva.valor_dia = (reserva.valor_hosp + reserva.valor_acomp1 + reserva.valor_acomp2 + reserva.valor_acomp3 + reserva.valor_acomp4 + reserva.valor_acomp5)
        else:
            # Se reserva.valor_hosp não for numérico, defina reserva.valor_dia como zero ou algum outro valor padrão
            reserva.valor_dia = 0
                   

        # print("Tipo de dado do valor:", type(preco_status_graduacao.valor))
        print("Valor do hospede:", reserva.valor_hosp)
        print("Valores recebidos do formulário:")
        print("Água:", novo_agua)
        print("Refri:", novo_refri)
        print("Cerveja:", novo_cerveja)
        reserva.qtde_acomp = reserva.qtde_hosp - 1
        
        # Verifica se os valores não são None antes de converter para Decimal
        if novo_agua is not None:
            reserva.qtde_agua = Decimal(novo_agua)
        if novo_refri is not None:
            reserva.qtde_refri = Decimal(novo_refri)
        if novo_cerveja is not None:
            reserva.qtde_cerveja = Decimal(novo_cerveja)
        if novo_pet is not None:
            reserva.qtde_pet = Decimal(novo_pet)    

        
            
        # Busca os valores dos produtos no banco de dados
        produto_agua = Produto.objects.get(nome='Água')
        produto_refri = Produto.objects.get(nome='Refrigerante')
        produto_cerveja = Produto.objects.get(nome='Cerveja')   
        produto_pet = Produto.objects.get(nome='Pet')  
        
        if reserva.motivo_viagem == "Saúde":
            reserva.desc_saude = reserva.valor_dia / 2
            reserva.subtotal = reserva.desc_saude * reserva.diarias          
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao  
            reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
            reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
            reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
            if reserva.qtde_pet != 0:
                reserva.total_pet = reserva.diarias * produto_pet.valor if reserva.diarias is not None else Decimal(0)
            else:
                reserva.total_pet = 0;    
            reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja 
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao + reserva.total_pet
        else:
            reserva.desc_saude = 0
            reserva.subtotal = reserva.valor_dia * reserva.diarias          
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao  
            reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
            reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
            reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
            if reserva.qtde_pet != 0:
                reserva.total_pet = reserva.diarias * reserva.qtde_pet * produto_pet.valor if reserva.diarias is not None else Decimal(0)
            else:
                reserva.total_pet = 0; 
            reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao + reserva.total_pet  


        # Calcula consumacao
        reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
        reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
        reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
        reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
        reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao + reserva.total_pet
        
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao_checkout'))
    
    return render(request, 'portal/consumacao.html', {'reserva': reserva})




@login_required
def relatorio_mensal(request):
    if request.method == 'GET':
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        print("Mês Relatório:", mes_relatorio)  # Verifique se os valores estão corretos
        print("Ano Relatório:", ano_relatorio)
        print("Forma Pagamento:", forma_pagamento)
        
        # Verificar se mes_relatorio e ano_relatorio não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio e ano_relatorio são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")
        
        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")
        
        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        )

        total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma'] 

          

        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        )

        total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma'] 

        # Atualização em lote dos objetos BaseDados
        if consultar_htm1.exists():
            consultar_htm1.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        if consultar_htm2.exists():
            consultar_htm2.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        context = {
            'consultar_htm1': consultar_htm1,
            'consultar_htm2': consultar_htm2,
            'total_htm1': total_htm1,
            'total_htm2': total_htm2,

        }
        return render(request, 'portal/relatorio_mensal.html', context)
    else:
        return HttpResponseBadRequest("Método de requisição inválido.")

@login_required
def relatorio_pagamento(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento.html', context) 

@login_required
def relatorio_pagamento_pix(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_pix.html', context) 

@login_required
def relatorio_pagamento_dinheiro(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_dinheiro.html', context) 





@login_required
def relatorio_pagamento_excel(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_excel.html', context) 



class RelatorioPix(View):
    nome_colunas_excel = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    cabecalho_fixo = 9
    cabecalho_dados1 = 1
    cabecalho_dados2 = 1
    linhas_dados1 = None
    linhas_dados2 = None
    mes_texto = None
    total_linhas = None
    total_htm1 = 0
    total_htm2 = 0
    dados1 = None
    dados2 = None
    mes_relatorio = None
    ano_relatorio = None


    def get(self, request, *args, **kwargs):
        self.mes_relatorio = request.GET.get('mes_relatorio')
        self.ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
        # print("Ano Relatório: auditoria", ano_relatorio)
        # print("Forma Pagamento: auditoria", forma_pagamento)

        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        
        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

        # Validação dos dados de entrada
        try:
            self.mes_relatorio = int(self.mes_relatorio)
            self.ano_relatorio = int(self.ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if self.mes_relatorio < 1 or self.mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
        print(self.mes_texto)

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        # Consulta para o segundo tipo de dados
        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        if not consultar_htm1.exists() and not consultar_htm2.exists():
            return render(request, 'portal/relatorio_pagamento_pix.html', {'exibir_modal': True})

        self.total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']  

        self.total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']

        # Combine os resultados das consultas em um único conjunto de dados
        self.dados1 = list(consultar_htm1)
        self.dados2 = list(consultar_htm2)

        self.linhas_dados1 = len(self.dados1) 
        self.linhas_dados2 = len(self.dados2)       
        self.total_linhas = self.linhas_dados1 + self.linhas_dados2

        # Gere a planilha
        workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

        # Configura o Excel
        self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

        # Configura o Estilo
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

        # Retorna a resposta HTTP com a planilha anexada
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

        # Salva o workbook como uma resposta HTTP
        workbook.save(response)
        return response

    def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'{ self.mes_texto}'

        # INCLUIR CABEÇALHO DADOS 1
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=10, column=col).value = coluna

        # INCLUIR DADOS 1
        self.escrever_dados(sheet, 11, self.dados1)

        # INCLUIR CABEÇALHO DADOS 2
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

        # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
        self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

        return workbook

    def escrever_dados(self, sheet, start_row, dados):
        # Escrever os dados
        for row, registro in enumerate(dados, start=start_row):
            for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
                valor = getattr(registro, coluna, None)
                if coluna == 'valor_total' and valor is not None:
                    try:
                        valor = float(valor)
                    except ValueError:
                        valor = None
                elif coluna == 'graduacao':
                # Verifica se o status é diferente de "MILITAR DA ATIVA" ou "MILITAR DA RESERVA"
                    if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
                        valor = "CIVIL"        
                sheet.cell(row=row, column=col).value = valor

    def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
        # Acessa a planilha ativa
        sheet = workbook.active

        # Larguras das colunas desejadas
        larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras_colunas.items():
            sheet.column_dimensions[col].width = width

        # Formatar coluna 'A' (Saida) para dd/mm/aaaa
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'

        # Aplicar estilo à coluna 'A' (Saida)
        for cell in sheet['A']:
            cell.style = date_format

        # Ocultar as linhas de grade
        sheet.sheet_view.showGridLines = False

        # Formatar coluna 'H' (valor_total) para ##.###,##
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

        # Aplicar estilo à coluna 'H' (valor_total)
        for cell in sheet['H']:
            cell.style = valor_format

        # Consulta para os dados do banco
        consultar_dados = self.dados1

        # Configura outros estilos, como bordas, fontes, etc.
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
        # Centralizar as colunas especificadas
        colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
        for coluna in colunas_centralizadas:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
         
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

    # def escrever_dados(self, sheet, start_row, dados):
    #     # Defina as colunas desejadas diretamente aqui
    #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

    #     # Escrever os dados
    #     for row, registro in enumerate(dados, start=start_row):
    #         for col, coluna in enumerate(colunas_desejadas, start=1):
    #             valor = getattr(registro, coluna, None)
    #             if coluna == 'valor_total' and valor is not None:
    #                 try:
    #                     valor = float(valor)
    #                 except ValueError:
    #                     valor = None
    #             sheet.cell(row=row, column=col).value = valor

    def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
        sheet = workbook.active
        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
        # total_linhas = linhas_dados1 + linhas_dados2

        
        # total_linhas = len(dados1) + len(dados2)
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        meses2 = {
            'JAN': 'janeiro',
            'FEV': 'fevereiro',
            'MAR': 'março',
            'ABR': 'abril',
            'MAI': 'maio',
            'JUN': 'junho',
            'JUL': 'julho',
            'AGO': 'agosto',
            'SET': 'setembro',
            'OUT': 'outubro',
            'NOV': 'novembro',
            'DEZ': 'dezembro'
        }

        self.mes2_texto = meses2.get(self.mes_texto)

        # Defina os estilos de borda
        dotted_border = Border(left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define as fontes
        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        # Define o preenchimento cinza claro para o cabeçalho
        gray_fill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')

        # Aplica os estilos para todas as células
        for row in sheet.iter_rows():
            for cell in row:
                # Aplica a borda
                cell.border = thin_border

                # Aplica a fonte padrão
                cell.font = fonte_padrao

        # Aplica estilos para o cabeçalho
        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # Aplica estilos para o cabeçalho 2
        for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # # Defina o estilo de borda externa
        # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                     top=Side(style='thin'), bottom=Side(style='thin'))
        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
        # for col in range(1, 9):
        #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
        #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
        # Seu texto
        texto = """MINISTERIO DA DEFESA
EXÉRCITO BRASILEIRO
COMANDO MILITAR DO OESTE
4ª BRIGADA DE CAVALARIA MECANIZADA
11º REGIMENTO DE CAVALARIA MECANIZADO
REGIMENTO MARECHAL DUTRA
"""

        # Insere o texto na célula A1
        sheet['A1'] = """MINISTERIO DA DEFESA"""
        sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
        sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
        sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
        sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
        sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
        sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
        sheet['A9'] = """HTM 1"""
        sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
        if total_htm1 is None:
            total_htm1 = decimal.Decimal('0.0')
        if total_htm2 is None:
            total_htm2 = decimal.Decimal('0.0')
        total_geral = total_htm1 + total_htm2
        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM PIX"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral

        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. PIX"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        # Define o idioma como português brasileiro
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        # Obter a data de hoje
        data_hoje = datetime.datetime.now()
        # Formatar a data como "dia de mês de ano"
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        # Inserir a data formatada na célula desejada
        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

        # Ajusta o alinhamento para envolver o texto e centralizá-lo
        sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

        # Mescla as células A1:H1
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

        # Define as células A7 e A9 com fonte negrito
        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



        # # tirando borda interna:

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        


# class RelatorioDinheiro(View):
#     nome_colunas_excel = {
#         'saida': 'Data de Saída',
#         'graduacao': 'PST/GRAD',
#         'nome_pagante': 'Hóspede',
#         'cpf_pagante': 'CPF',
#         'qtde_acomp': 'Nr Acomp',
#         'uh': 'UH',
#         'diarias': 'Dias',
#         'valor_total': 'Total (R$)',
#     }

#     cabecalho_fixo = 9
#     cabecalho_dados1 = 1
#     cabecalho_dados2 = 1
#     linhas_dados1 = None
#     linhas_dados2 = None
#     mes_texto = None
#     total_linhas = None
#     total_htm1 = 0
#     total_htm2 = 0
#     dados1 = None
#     dados2 = None
#     mes_relatorio = None
#     ano_relatorio = None


#     def get(self, request, *args, **kwargs):
#         self.mes_relatorio = request.GET.get('mes_relatorio')
#         self.ano_relatorio = request.GET.get('ano_relatorio')
#         forma_pagamento = request.GET.get('forma_pagamento')
#         # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
#         # print("Ano Relatório: auditoria", ano_relatorio)
#         # print("Forma Pagamento: auditoria", forma_pagamento)

#         meses = {
#             '1': 'JAN',
#             '2': 'FEV',
#             '3': 'MAR',
#             '4': 'ABR',
#             '5': 'MAI',
#             '6': 'JUN',
#             '7': 'JUL',
#             '8': 'AGO',
#             '9': 'SET',
#             '10': 'OUT',
#             '11': 'NOV',
#             '12': 'DEZ'
#         }

        
#         # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
#         if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
#             return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

#         # Validação dos dados de entrada
#         try:
#             self.mes_relatorio = int(self.mes_relatorio)
#             self.ano_relatorio = int(self.ano_relatorio)
#         except ValueError:
#             return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

#         if self.mes_relatorio < 1 or self.mes_relatorio > 12:
#             return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

#         if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
#             return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
#         self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
#         print(self.mes_texto)

#         # Realize as consultas usando os parâmetros recebidos
#         consultar_htm1 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_01",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).order_by('saida')

#         # Consulta para o segundo tipo de dados
#         consultar_htm2 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_02",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).order_by('saida')

#         if not consultar_htm1.exists() and not consultar_htm2.exists():
#             return render(request, 'portal/relatorio_pagamento_dinheiro.html', {'exibir_modal': True})

#         self.total_htm1 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_01",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).aggregate(soma=Sum('valor_total'))['soma']  

#         self.total_htm2 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_02",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).aggregate(soma=Sum('valor_total'))['soma']

#         # Combine os resultados das consultas em um único conjunto de dados
#         self.dados1 = list(consultar_htm1)
#         self.dados2 = list(consultar_htm2)

#         self.linhas_dados1 = len(self.dados1) 
#         self.linhas_dados2 = len(self.dados2)       
#         self.total_linhas = self.linhas_dados1 + self.linhas_dados2

#         # Gere a planilha
#         workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

#         # Configura o Excel
#         self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

#         # Configura o Estilo
#         self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

#         # Retorna a resposta HTTP com a planilha anexada
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

#         # Salva o workbook como uma resposta HTTP
#         workbook.save(response)
#         return response

#     def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
#         workbook = Workbook()
#         sheet = workbook.active
#         sheet.title = f'{ self.mes_texto}'

#         # INCLUIR CABEÇALHO DADOS 1
#         for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
#             sheet.cell(row=10, column=col).value = coluna

#         # INCLUIR DADOS 1
#         self.escrever_dados(sheet, 11, self.dados1)

#         # INCLUIR CABEÇALHO DADOS 2
#         for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
#             sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

#         # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
#         self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

#         return workbook

#     def escrever_dados(self, sheet, start_row, dados):
#         # Escrever os dados
#         for row, registro in enumerate(dados, start=start_row):
#             for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
#                 valor = getattr(registro, coluna, None)
#                 if coluna == 'valor_total' and valor is not None:
#                     try:
#                         valor = float(valor)
#                     except ValueError:
#                         valor = None
#                 elif coluna == 'graduacao':
#                 # Verifica se o status é diferente de "MILITAR DA ATIVA" ou "MILITAR DA RESERVA"
#                     if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
#                         valor = "CIVIL"        
#                 sheet.cell(row=row, column=col).value = valor

#     def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
#         # Acessa a planilha ativa
#         sheet = workbook.active

#         # Larguras das colunas desejadas
#         larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
#         for col, width in larguras_colunas.items():
#             sheet.column_dimensions[col].width = width

#         # Formatar coluna 'A' (Saida) para dd/mm/aaaa
#         date_format = NamedStyle(name='date_format')
#         date_format.number_format = 'DD/MM/YYYY'

#         # Aplicar estilo à coluna 'A' (Saida)
#         for cell in sheet['A']:
#             cell.style = date_format

#         # Ocultar as linhas de grade
#         sheet.sheet_view.showGridLines = False

#         # Formatar coluna 'H' (valor_total) para ##.###,##
#         valor_format = NamedStyle(name='valor_format')
#         valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

#         # Aplicar estilo à coluna 'H' (valor_total)
#         for cell in sheet['H']:
#             cell.style = valor_format

#         # Consulta para os dados do banco
#         consultar_dados = self.dados1

#         # Configura outros estilos, como bordas, fontes, etc.
#         self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
#         # Centralizar as colunas especificadas
#         colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
#         for coluna in colunas_centralizadas:
#             for cell in sheet[coluna]:
#                 cell.alignment = Alignment(horizontal='center')

#         # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
#         # dados1 = list(consultar_htm1)
#         # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
#         # dados2 = list(consultar_htm2)
#         # linhas_dados1 = len(dados1) 
#         # linhas_dados2 = len(dados2)       
         
#         sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
#         sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
#         sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

#     # def escrever_dados(self, sheet, start_row, dados):
#     #     # Defina as colunas desejadas diretamente aqui
#     #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

#     #     # Escrever os dados
#     #     for row, registro in enumerate(dados, start=start_row):
#     #         for col, coluna in enumerate(colunas_desejadas, start=1):
#     #             valor = getattr(registro, coluna, None)
#     #             if coluna == 'valor_total' and valor is not None:
#     #                 try:
#     #                     valor = float(valor)
#     #                 except ValueError:
#     #                     valor = None
#     #             sheet.cell(row=row, column=col).value = valor

#     def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
#         sheet = workbook.active
#         # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
#         # dados1 = list(consultar_htm1)
#         # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
#         # dados2 = list(consultar_htm2)
        
#         # linhas_dados1 = len(dados1) 
#         # linhas_dados2 = len(dados2)       
#         # total_linhas = linhas_dados1 + linhas_dados2

        
#         # total_linhas = len(dados1) + len(dados2)
#         sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
#         meses2 = {
#             'JAN': 'janeiro',
#             'FEV': 'fevereiro',
#             'MAR': 'março',
#             'ABR': 'abril',
#             'MAI': 'maio',
#             'JUN': 'junho',
#             'JUL': 'julho',
#             'AGO': 'agosto',
#             'SET': 'setembro',
#             'OUT': 'outubro',
#             'NOV': 'novembro',
#             'DEZ': 'dezembro'
#         }

#         self.mes2_texto = meses2.get(self.mes_texto)

#         # Defina os estilos de borda
#         dotted_border = Border(left=Side(style='dotted'),
#                             right=Side(style='dotted'),
#                             top=Side(style='dotted'),
#                             bottom=Side(style='dotted'))

#         thin_border = Border(left=Side(style='thin'),
#                             right=Side(style='thin'),
#                             top=Side(style='thin'),
#                             bottom=Side(style='thin'))

#         # Define as fontes
#         fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
#         fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

#         # Define o preenchimento cinza claro para o cabeçalho
#         gray_fill = PatternFill(start_color='00CCCCCC',
#                                 end_color='00CCCCCC',
#                                 fill_type='solid')

#         # Aplica os estilos para todas as células
#         for row in sheet.iter_rows():
#             for cell in row:
#                 # Aplica a borda
#                 cell.border = thin_border

#                 # Aplica a fonte padrão
#                 cell.font = fonte_padrao

#         # Aplica estilos para o cabeçalho
#         for row in sheet.iter_rows(min_row=10, max_row=10):
#             for cell in row:
#                 cell.font = fonte_negrito
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 cell.fill = gray_fill

#         # Aplica estilos para o cabeçalho 2
#         for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
#             for cell in row:
#                 cell.font = fonte_negrito
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 cell.fill = gray_fill

#         # # Defina o estilo de borda externa
#         # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#         #                     top=Side(style='thin'), bottom=Side(style='thin'))
#         # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
#         # for col in range(1, 9):
#         #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
#         #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
#         # Seu texto
#         texto = """MINISTERIO DA DEFESA
# EXÉRCITO BRASILEIRO
# COMANDO MILITAR DO OESTE
# 4ª BRIGADA DE CAVALARIA MECANIZADA
# 11º REGIMENTO DE CAVALARIA MECANIZADO
# REGIMENTO MARECHAL DUTRA
# """

#         # Insere o texto na célula A1
#         sheet['A1'] = """MINISTERIO DA DEFESA"""
#         sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
#         sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
#         sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
#         sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
#         sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
#         sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
#         sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
#         sheet['A9'] = """HTM 1"""
#         sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
#         sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
#         sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
#         total_htm1_decimal = decimal.Decimal(total_htm1 or 0)
#         total_htm2_decimal = decimal.Decimal(total_htm2 or 0)

#         total_geral = total_htm1_decimal + total_htm2_decimal
#         print("Valor de total_htm1:", total_htm1)
#         print("Valor de total_htm2:", total_htm2)
#         sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM DINHEIRO"
#         sheet.cell(row=total_linhas + 15, column=8).value = total_geral

#         sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
#         sheet.cell(row=total_linhas + 18, column=1).value = "DEP. DINHEIRO"
#         sheet.cell(row=total_linhas + 18, column=8).value = total_geral

#         # Define o idioma como português brasileiro
#         locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
#         # Obter a data de hoje
#         data_hoje = datetime.now()
#         # Formatar a data como "dia de mês de ano"
#         data_formatada = data_hoje.strftime("%d de %B de %Y")

#         # Inserir a data formatada na célula desejada
#         sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
#         sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
#         sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
#         sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
#         sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

#         # Ajusta o alinhamento para envolver o texto e centralizá-lo
#         sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
#         sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

#         # Mescla as células A1:H1
#         sheet.merge_cells('A1:H1')
#         sheet.merge_cells('A2:H2')
#         sheet.merge_cells('A3:H3')
#         sheet.merge_cells('A4:H4')
#         sheet.merge_cells('A5:H5')
#         sheet.merge_cells('A6:H6')
#         sheet.merge_cells('A7:H7')
#         sheet.merge_cells('A8:H8')
#         sheet.merge_cells('A9:H9')
#         sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
#         sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
#         sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
#         sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

#         sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

#         # Define as células A7 e A9 com fonte negrito
#         sheet['A7'].font = fonte_negrito
#         sheet['A9'].font = fonte_negrito
#         sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
#         sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
#         sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
#         sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


#         sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
#         sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
#         sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
#         sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
#         sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
#         sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
#         sheet.cell(row=total_linhas + 14, column=8).border = thin_border
#         sheet.cell(row=total_linhas + 15, column=8).border = thin_border
#         sheet.cell(row=total_linhas + 17, column=1).border = thin_border
#         sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 8):
#         #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 9):
#         #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 8):
#         #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 8):
#         #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

#         # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
#         # for col in range(1, 9):
#         #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



#         # # tirando borda interna:

#         for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
#             for cell in row:
#                 if cell.row == 1:
#                     cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
#                 elif cell.row == 6:
#                     cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
#                 else:
#                     cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

#         for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
#             for cell in row:
#                 cell.border = None

#         for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
#             for cell in row:
#                 cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        






class RelatorioDinheiro(View):
    nome_colunas_excel = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    cabecalho_fixo = 9
    cabecalho_dados1 = 1
    cabecalho_dados2 = 1
    linhas_dados1 = None
    linhas_dados2 = None
    mes_texto = None
    total_linhas = None
    total_htm1 = 0
    total_htm2 = 0
    dados1 = None
    dados2 = None
    mes_relatorio = None
    ano_relatorio = None


    def get(self, request, *args, **kwargs):
        self.mes_relatorio = request.GET.get('mes_relatorio')
        self.ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
        # print("Ano Relatório: auditoria", ano_relatorio)
        # print("Forma Pagamento: auditoria", forma_pagamento)

        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        
        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

        # Validação dos dados de entrada
        try:
            self.mes_relatorio = int(self.mes_relatorio)
            self.ano_relatorio = int(self.ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if self.mes_relatorio < 1 or self.mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
        print(self.mes_texto)

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        # Consulta para o segundo tipo de dados
        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        if not consultar_htm1.exists() and not consultar_htm2.exists():
            return HttpResponse("Não há lançamentos para o mês solicitado.")
        


        self.total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']  

        self.total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']

        # if not self.total_htm1.exists() and not self.total_htm2.exists():
        #     return HttpResponse("Não há lançamentos para o mês solicitado.")

        # Combine os resultados das consultas em um único conjunto de dados
        self.dados1 = list(consultar_htm1)
        self.dados2 = list(consultar_htm2)

        self.linhas_dados1 = len(self.dados1) 
        self.linhas_dados2 = len(self.dados2)       
        self.total_linhas = self.linhas_dados1 + self.linhas_dados2

        # Gere a planilha
        workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

        # Configura o Excel
        self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

        # Configura o Estilo
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

        # Retorna a resposta HTTP com a planilha anexada
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

        # Salva o workbook como uma resposta HTTP
        workbook.save(response)
        return response

    def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'{ self.mes_texto}'

        # INCLUIR CABEÇALHO DADOS 1
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=10, column=col).value = coluna

        # INCLUIR DADOS 1
        self.escrever_dados(sheet, 11, self.dados1)

        # INCLUIR CABEÇALHO DADOS 2
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

        # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
        self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

        return workbook

    def escrever_dados(self, sheet, start_row, dados):
        # Escrever os dados
        for row, registro in enumerate(dados, start=start_row):
            for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
                valor = getattr(registro, coluna, None)
                if coluna == 'valor_total' and valor is not None:
                    try:
                        valor = float(valor)
                    except ValueError:
                        valor = None
                elif coluna == 'graduacao':
                # Verifica se o status é diferente de "MILITAR DA ATIVA" ou "MILITAR DA RESERVA"
                    if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
                        valor = "CIVIL"         
                sheet.cell(row=row, column=col).value = valor

    def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
        # Acessa a planilha ativa
        sheet = workbook.active

        # Larguras das colunas desejadas
        larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras_colunas.items():
            sheet.column_dimensions[col].width = width

        # Formatar coluna 'A' (Saida) para dd/mm/aaaa
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'

        # Aplicar estilo à coluna 'A' (Saida)
        for cell in sheet['A']:
            cell.style = date_format

        # Ocultar as linhas de grade
        sheet.sheet_view.showGridLines = False

        # Formatar coluna 'H' (valor_total) para ##.###,##
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

        # Aplicar estilo à coluna 'H' (valor_total)
        for cell in sheet['H']:
            cell.style = valor_format

        # Consulta para os dados do banco
        consultar_dados = self.dados1

        # Configura outros estilos, como bordas, fontes, etc.
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
        # Centralizar as colunas especificadas
        colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
        for coluna in colunas_centralizadas:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
         
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

    # def escrever_dados(self, sheet, start_row, dados):
    #     # Defina as colunas desejadas diretamente aqui
    #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

    #     # Escrever os dados
    #     for row, registro in enumerate(dados, start=start_row):
    #         for col, coluna in enumerate(colunas_desejadas, start=1):
    #             valor = getattr(registro, coluna, None)
    #             if coluna == 'valor_total' and valor is not None:
    #                 try:
    #                     valor = float(valor)
    #                 except ValueError:
    #                     valor = None
    #             sheet.cell(row=row, column=col).value = valor

    def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
        sheet = workbook.active
        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
        # total_linhas = linhas_dados1 + linhas_dados2

        
        # total_linhas = len(dados1) + len(dados2)
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        meses2 = {
            'JAN': 'janeiro',
            'FEV': 'fevereiro',
            'MAR': 'março',
            'ABR': 'abril',
            'MAI': 'maio',
            'JUN': 'junho',
            'JUL': 'julho',
            'AGO': 'agosto',
            'SET': 'setembro',
            'OUT': 'outubro',
            'NOV': 'novembro',
            'DEZ': 'dezembro'
        }

        self.mes2_texto = meses2.get(self.mes_texto)

        # Defina os estilos de borda
        dotted_border = Border(left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define as fontes
        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        # Define o preenchimento cinza claro para o cabeçalho
        gray_fill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')

        # Aplica os estilos para todas as células
        for row in sheet.iter_rows():
            for cell in row:
                # Aplica a borda
                cell.border = thin_border

                # Aplica a fonte padrão
                cell.font = fonte_padrao

        # Aplica estilos para o cabeçalho
        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # Aplica estilos para o cabeçalho 2
        for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # # Defina o estilo de borda externa
        # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                     top=Side(style='thin'), bottom=Side(style='thin'))
        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
        # for col in range(1, 9):
        #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
        #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
        # Seu texto
        texto = """MINISTERIO DA DEFESA
EXÉRCITO BRASILEIRO
COMANDO MILITAR DO OESTE
4ª BRIGADA DE CAVALARIA MECANIZADA
11º REGIMENTO DE CAVALARIA MECANIZADO
REGIMENTO MARECHAL DUTRA
"""

        # Insere o texto na célula A1
        sheet['A1'] = """MINISTERIO DA DEFESA"""
        sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
        sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
        sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
        sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
        sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
        sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
        sheet['A9'] = """HTM 1"""
        sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
        total_htm1_decimal = decimal.Decimal(total_htm1 or 0)
        total_htm2_decimal = decimal.Decimal(total_htm2 or 0)
        total_geral = total_htm1_decimal + total_htm2_decimal
        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM DINHEIRO"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral

        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. DINHEIRO"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        # Define o idioma como português brasileiro
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        # Obter a data de hoje
        data_hoje = datetime.datetime.now()
        # Formatar a data como "dia de mês de ano"
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        # Inserir a data formatada na célula desejada
        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

        # Ajusta o alinhamento para envolver o texto e centralizá-lo
        sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

        # Mescla as células A1:H1
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

        # Define as células A7 e A9 com fonte negrito
        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



        # # tirando borda interna:

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        





# def obter_ocupacao_hotel(mes, ano, hotel):
#     ocupacao_hotel = {}
#     dias_no_mes = monthrange(ano, mes)[1]
    
#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}
#         for quarto in range(1, 13 if hotel == 'HTM_01' else 4):  # Determinar o número de quartos com base no hotel
#             # Verificar se há reservas para este dia e quarto
#             reservas = BaseDados.objects.filter(
#                 entrada__year=ano,
#                 entrada__month=mes,
#                 saida__year=ano,
#                 saida__month=mes,
#                 mhex=hotel,
#                 uh=str(quarto),
#                 entrada__day__lte=dia,
#                 saida__day__gte=dia
#             )
#             ocupacao_hotel[dia][quarto] = reservas.exists()
    
#     return ocupacao_hotel

# def visualizar_ocupacao_hotel(request):
#     mes = 3  # Janeiro
#     ano = 2024  # Ano desejado
#     ocupacao_htm_01 = obter_ocupacao_hotel(mes, ano, 'HTM_01')
#     ocupacao_htm_02 = obter_ocupacao_hotel(mes, ano, 'HTM_02')
    
#     return render(request, 'portal/ocupacao.html', {
#         'ocupacao_htm_01': ocupacao_htm_01,
#         'ocupacao_htm_02': ocupacao_htm_02,
#     })



# def obter_ocupacao_hotel(mes, ano, hotel):
#     ocupacao_hotel = {}
#     dias_no_mes = monthrange(ano, mes)[1]

#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}

#         for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
#             reservas_por_dia = BaseDados.objects.filter(
#                 saida__year=ano,
#                 saida__month=mes,
#                 mhex=hotel,
#                 saida__day=dia,  # Considerar apenas reservas com data de saída igual ao dia atual
#                 uh=str(quarto)  # Filtrar pelo número do quarto
#             )

#             if reservas_por_dia.exists():
#                 for reserva in reservas_por_dia:
#                     diarias = reserva.diarias
#                     valor_total = reserva.valor_total
#                     media_diaria = valor_total / diarias if diarias > 0 else 0
#                     for i in range(dia - diarias + 1, dia + 1):
#                         if i > 0:
#                             ocupacao_hotel[i][quarto] = {
#                                 'ocupado': True,
#                                 'media_diaria': "{:.2f}".format(media_diaria)  # Formatação decimal com duas casas decimais
#                             }
#             else:
#                 ocupacao_hotel[dia][quarto] = {
#                     'ocupado': False,
#                     'media_diaria': ""  # Deixando a string vazia para indicar que não há ocupação
#                 }

#     return ocupacao_hotel

######################################################################################################################################

# from decimal import Decimal
# from django.db.models import F, Sum, ExpressionWrapper, DecimalField
# from calendar import monthrange
# from datetime import datetime, timedelta


# def obter_ocupacao_hotel(mes, ano, hotel):
#     ocupacao_hotel = {}
#     dias_no_mes = monthrange(ano, mes)[1]
#     primeiro_dia_mes = datetime(ano, mes, 1).date()  # Primeiro dia do mês selecionado
#     ultimo_dia_mes = datetime(ano, mes, dias_no_mes).date()  # Último dia do mês selecionado

#     # Inicializar o dicionário para todos os dias e quartos
#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}
#         for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
#             ocupacao_hotel[dia][quarto] = {
#                 'ocupado': False,
#                 'diaria': ""  # Inicialmente vazio, indicando que não há reserva
#             }

#     # Consultar todas as reservas do hotel, independente do mês
#     reservas = BaseDados.objects.filter(
#         mhex=hotel,
#         status_reserva="Pago"
#     )

#     # Processar as reservas para calcular os dias de ocupação com base na entrada, saída e número de diárias
#     for reserva in reservas:
#         entrada = reserva.entrada  # Data de entrada
#         saida = reserva.saida  # Data de saída
        
#         # Verificar se entrada ou saída são None
#         if entrada is None or saida is None:
#             continue  # Pule esta reserva se algum desses campos estiver ausente
        
#         saida = saida - timedelta(days=1)  # Excluir o dia de saída
#         diarias = (saida - entrada).days + 1  # Número total de diárias da reserva
#         quarto = int(reserva.uh)
#         valor_dia = reserva.valor_dia  # Valor da diária 

#         # Calcular o valor adicional de pets (TOTAL_PET dividido por QTDE_PET e pelo número de dias da reserva)
#         total_pet = reserva.total_pet  # O valor de total_pet é garantido como Decimal pelo modelo
#         qtde_pet = reserva.qtde_pet if reserva.qtde_pet is not None else 0  # Garantir que qtde_pet não seja None
#         valor_pet_diaria = Decimal(0)  # Inicializamos como zero

#         if qtde_pet > 0:  # Evitar divisão por zero
#             valor_pet_diaria = total_pet / (qtde_pet * diarias)  # Valor proporcional por diária para pets

#         # Verificar se a reserva impacta o mês selecionado
#         if entrada <= ultimo_dia_mes and saida >= primeiro_dia_mes:
#             # Ajustar o início da ocupação para o mês selecionado, se a entrada for antes
#             inicio_ocupacao = max(entrada, primeiro_dia_mes)
#             fim_ocupacao = min(saida, ultimo_dia_mes)

#             # Marcar os dias de ocupação entre o início e o fim da ocupação ajustados
#             for dia_ocupacao in range(inicio_ocupacao.day, fim_ocupacao.day + 1):
#                 ocupacao_hotel[dia_ocupacao][quarto]['ocupado'] = True
                
#                 # Se já existe um valor na diária, somar com o novo valor
#                 if isinstance(ocupacao_hotel[dia_ocupacao][quarto]['diaria'], Decimal):
#                     ocupacao_hotel[dia_ocupacao][quarto]['diaria'] += valor_dia + valor_pet_diaria
#                 else:
#                     ocupacao_hotel[dia_ocupacao][quarto]['diaria'] = valor_dia + valor_pet_diaria

#     return ocupacao_hotel

from decimal import Decimal, ROUND_HALF_UP
from calendar import monthrange
from datetime import datetime, timedelta

def obter_ocupacao_hotel(mes, ano, hotel):
    # Inicialização do dicionário de ocupação
    ocupacao_hotel = {}
    dias_no_mes = monthrange(ano, mes)[1]
    primeiro_dia_mes = datetime.datetime(ano, mes, 1).date()  # Usando datetime.datetime
    ultimo_dia_mes = datetime.datetime(ano, mes, dias_no_mes).date()  # Usando datetime.datetime

    # Criação da estrutura inicial para os quartos e dias
    for dia in range(1, dias_no_mes + 1):
        ocupacao_hotel[dia] = {}
        for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
            ocupacao_hotel[dia][quarto] = {
                'ocupado': False,
                'diaria': Decimal('0.00'),
                'reserva_id': None  # Inicializado como None para rastrear a reserva
            }

    # Consultar todas as reservas do hotel
    reservas = BaseDados.objects.filter(
        mhex=hotel,
        status_reserva="Pago"
    )

    # Processar as reservas
    for reserva in reservas:
        entrada = reserva.entrada
        saida = reserva.saida
        if entrada is None or saida is None:
            continue  # Pula reservas com dados incompletos

        # Ajusta saída para não incluir o dia de saída
        saida = saida - timedelta(days=1)

        # Calcula o número de diárias
        diarias = (saida - entrada).days + 1

        # Certifica que o número de diárias seja válido
        if diarias <= 0:
            continue

        # Extrai informações da reserva
        quarto = int(reserva.uh)
        valor_total = Decimal(str(reserva.valor_total))
        valor_dia = (valor_total / diarias).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

        # Verifica se a reserva impacta o mês e o ano selecionados
        if (entrada.month == mes and entrada.year == ano) or (saida.month == mes and saida.year == ano):
            inicio_ocupacao = max(entrada, primeiro_dia_mes)
            fim_ocupacao = min(saida, ultimo_dia_mes)

            # Marcar ocupação diária no período da reserva
            for dia_ocupacao in range(inicio_ocupacao.day, fim_ocupacao.day + 1):
                if ocupacao_hotel[dia_ocupacao][quarto]['ocupado']:
                    # Se o quarto já está ocupado, exibir alerta com ID da reserva existente e atual
                    print(
                        f"Alerta: Sobreposição detectada! Dia: {dia_ocupacao}, Quarto: {quarto}, "
                        f"Reserva existente: {ocupacao_hotel[dia_ocupacao][quarto]['reserva_id']}, "
                        f"Reserva atual: {reserva.id}, "
                        f"Diária existente: {ocupacao_hotel[dia_ocupacao][quarto]['diaria']}, "
                        f"Diária atual: {valor_dia}"
                    )
                else:
                    # Marca como ocupado e registra o ID da reserva atual
                    ocupacao_hotel[dia_ocupacao][quarto]['ocupado'] = True
                    ocupacao_hotel[dia_ocupacao][quarto]['diaria'] = valor_dia
                    ocupacao_hotel[dia_ocupacao][quarto]['reserva_id'] = reserva.id

    # Garantir que quartos não ocupados tenham valor zero
    for dia in ocupacao_hotel:
        for quarto in ocupacao_hotel[dia]:
            if not ocupacao_hotel[dia][quarto]['ocupado']:
                ocupacao_hotel[dia][quarto]['diaria'] = Decimal('0.00')
                ocupacao_hotel[dia][quarto]['reserva_id'] = None  # Mantém reserva_id como None para quartos desocupados

    return ocupacao_hotel












class Visualizar_Ocupacao_Hotel(View):
    def get(self, request, *args, **kwargs):
        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')

        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes e ano são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        mes_texto = meses.get(str(mes_relatorio), 'DESC')
        

        ocupacao_htm_01 = obter_ocupacao_hotel(mes_relatorio, ano_relatorio, 'HTM_01')
        ocupacao_htm_02 = obter_ocupacao_hotel(mes_relatorio, ano_relatorio, 'HTM_02')

        return render(request, 'portal/ocupacao_comvalores.html', {
            'ocupacao_htm_01': ocupacao_htm_01,
            'ocupacao_htm_02': ocupacao_htm_02,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,
            'mes_texto': mes_texto
        })

@login_required
def relatorio_ocupacao_finaceiro(request):    
    return render(request, 'portal/relatorio_ocupacao_financeira.html')         


#############################################################################

def obter_censo_mes(mes, ano, hotel):
    relatorio = []
    dias_no_mes = monthrange(ano, mes)[1]

    # Inicializar o dicionário para todos os dias com campos QTDE e VALOR
    for dia in range(1, dias_no_mes + 1):
        relatorio.append({
            'dia': dia,
            'qtde': 0,  # Inicializar a quantidade
            'valor': 0.0  # Inicializar o valor total do dia
        })

    # Processar reservas do hotel específico (HTM_01 ou HTM_02)
    reservas = BaseDados.objects.filter(
        mhex=hotel,  # Filtrar para o hotel específico
        status_reserva="Pago"
    )

    # Iterar sobre cada dia do mês
    for dia in range(1, dias_no_mes + 1):
        # Garantir que data_atual seja um objeto datetime
        data_atual = datetime.datetime(ano, mes, dia)

        # Verificar as reservas ativas no dia atual (considerando entrada e saída)
        reservas_dia = reservas.filter(entrada__lte=data_atual, saida__gt=data_atual)

        # Para cada reserva do dia, acumular qtde e valor
        for reserva in reservas_dia:
            qtde_hosp = reserva.qtde_hosp or 0  # Usar o campo QTDE_HOSP para calcular a quantidade de hóspedes
            valor_dia = float(reserva.valor_dia)  # Usar diretamente o valor diário do banco de dados

            # Somar a quantidade de hóspedes e o valor para o dia
            relatorio[dia - 1]['qtde'] += qtde_hosp
            relatorio[dia - 1]['valor'] += valor_dia

            # Adicionar log detalhado
            print(f"[MENSAL] Dia {dia} - Reserva ID: {reserva.id}, Valor Dia: {valor_dia}, "
                f"Qtde Hospedes: {qtde_hosp}, Total Dia: {relatorio[dia - 1]['valor']}")

    return relatorio



class Visualizar_Censo_Mes(View):
    def get(self, request, *args, **kwargs):
        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        # Verificar se os parâmetros mes_relatorio e ano_relatorio estão presentes
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')

        if not mes_relatorio or not ano_relatorio:
            return render(request, 'portal/relatorio_censo_mes.html', {
                'error_message': "Por favor, selecione o mês e o ano para gerar o relatório."
            })

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")
        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Definir o texto do mês
        mes_texto = meses.get(str(mes_relatorio), 'DESC')

        # Obter dados de censo para HTM_01 e HTM_02 separadamente
        print(f"[DEBUG] Gerando censo para Mês: {mes_relatorio}, Ano: {ano_relatorio}, Hotel: HTM_01")
        censo_htm_01 = obter_censo_mes(mes_relatorio, ano_relatorio, 'HTM_01')

        # Log detalhado do censo de HTM_01
        for dia in censo_htm_01:
            print(f"[DEBUG HTM_01] Dia {dia['dia']}: Qtde: {dia['qtde']}, Valor: {dia['valor']}")

        print(f"[DEBUG] Gerando censo para Mês: {mes_relatorio}, Ano: {ano_relatorio}, Hotel: HTM_02")
        censo_htm_02 = obter_censo_mes(mes_relatorio, ano_relatorio, 'HTM_02')

        # Log detalhado do censo de HTM_02
        for dia in censo_htm_02:
            print(f"[DEBUG HTM_02] Dia {dia['dia']}: Qtde: {dia['qtde']}, Valor: {dia['valor']}")

        # Renderizar o template com os censos separados para cada hotel
        return render(request, 'portal/relatorio_censo_mes.html', {
            'censo_htm_01': censo_htm_01,
            'censo_htm_02': censo_htm_02,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,
            'mes_texto': mes_texto
        })


@login_required
def relatorio_censo_mes(request):    
    return render(request, 'portal/relatorio_censo_mes.html')   

#############################################################################

import tempfile
from openpyxl import Workbook
from django.http import FileResponse, HttpResponse
from openpyxl.styles import Alignment

def obter_censo_ano_excel(ano, hotel):
    # Criar o workbook e configurar a planilha
    wb = Workbook()
    ws = wb.active
    ws.title = f"Relatório {ano}"

    # Configurar cabeçalhos
    meses_nome = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 
                  'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
    ws.merge_cells('A1:A2')
    ws['A1'] = 'DIA'
    col = 2
    for mes in meses_nome:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        ws.cell(row=1, column=col, value=mes).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col, value='QTDE')
        ws.cell(row=2, column=col+1, value='VALOR')
        col += 2

    # Obter os dados do censo anual
    relatorio = obter_censo_ano(ano, hotel)

    # Preencher os dados na planilha
    for dia in range(1, 32):
        ws.cell(row=dia + 2, column=1, value=dia)  # Preencher a coluna dos dias
        col = 2
        for mes in range(1, 13):
            if dia <= len(relatorio[mes]):
                qtde = relatorio[mes][dia - 1]['qtde']
                valor = relatorio[mes][dia - 1]['valor']
                ws.cell(row=dia + 2, column=col, value=qtde)
                ws.cell(row=dia + 2, column=col+1, value=valor).number_format = '#,##0.00'
            else:
                ws.cell(row=dia + 2, column=col, value='-')
                ws.cell(row=dia + 2, column=col+1, value='-')
            col += 2

    # Salvar o arquivo Excel em um arquivo temporário
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)  # Garantir que o ponteiro do arquivo esteja no início
        response = FileResponse(open(tmp_file.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{hotel}_{ano}.xlsx"'
        return response

from django.http import HttpResponse
from django.views import View

class Visualizar_Censo_Ano_Excel(View):
    def get(self, request, *args, **kwargs):
        # Verificar se o parâmetro ano_relatorio está presente
        ano_relatorio = request.GET.get('ano_relatorio')

        if not ano_relatorio:
            return HttpResponse("Ano do relatório não foi fornecido.", status=400)

        # Validação do ano
        try:
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponse("Ano deve ser um número inteiro.", status=400)

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponse("Ano deve estar entre 2024 e 2030.", status=400)

        # Obter o hotel, aqui vamos assumir que está sendo passado via query params, ou fixar para testes
        hotel = request.GET.get('hotel', 'HTM_01')  # Default para HTM_01, você pode ajustar isso

        # Gerar o Excel para o hotel e ano
        return obter_censo_ano_excel(ano_relatorio, hotel)

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment

#############################################################################

from datetime import date, timedelta
from calendar import monthrange
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import BaseDados


def obter_censo_ano(ano, hotel):
    # Inicializar o relatório para armazenar os dados
    relatorio = {mes: [{'dia': dia, 'qtde': 0, 'valor': 0.0} for dia in range(1, monthrange(ano, mes)[1] + 1)] for mes in range(1, 13)}

    # Buscar reservas válidas no banco de dados
    reservas = BaseDados.objects.filter(
        mhex=hotel,
        status_reserva="Pago",
        entrada__year__lte=ano,
        saida__year__gte=ano
    )

    # Processar cada reserva
    for reserva in reservas:
        if not reserva.entrada or not reserva.saida or not reserva.valor_total:
            # Ignorar reservas inválidas
            continue

        # Calcular o número de dias da reserva
        dias_reserva = (reserva.saida - reserva.entrada).days
        if dias_reserva <= 0:
            # Ignorar reservas com duração inválida
            continue

        # Calcular o valor diário proporcional
        valor_diario = float(reserva.valor_total) // dias_reserva  # Divisão inteira
        sobra = float(reserva.valor_total) % dias_reserva  # Sobra do valor total
        qtde_hospedes = reserva.qtde_hosp or 0

        # Iterar pelos dias da reserva
        data_atual = reserva.entrada
        while data_atual < reserva.saida:
            if data_atual.year == ano:  # Garantir que estamos no ano correto
                mes = data_atual.month
                dia = data_atual.day

                # Adicionar sobra ao último dia da reserva
                if data_atual == reserva.saida - timedelta(days=1):  # Último dia
                    valor_diario += sobra

                # Adicionar ao relatório
                relatorio[mes][dia - 1]['qtde'] += qtde_hospedes
                relatorio[mes][dia - 1]['valor'] += valor_diario

            data_atual += timedelta(days=1)

    # Garantir que dias sem reservas sejam zerados
    for mes in relatorio:
        for dia in relatorio[mes]:
            if dia['valor'] == 0:
                dia['qtde'] = 0  # Zerar a quantidade de hóspedes se o valor for 0

    return relatorio



import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import BaseDados

class Visualizar_Censo_Ano(View):
    def get(self, request, *args, **kwargs):
        ano_relatorio = request.GET.get('ano_relatorio')

        if not ano_relatorio:
            return render(request, 'portal/relatorio_censo_ano.html', {
                'error_message': "Por favor, selecione o ano para gerar o relatório."
            })

        try:
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Ano deve ser um número inteiro.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Gerar relatórios para os hotéis HTM_01 e HTM_02
        censo_htm_01 = obter_censo_ano(ano_relatorio, 'HTM_01')
        censo_htm_02 = obter_censo_ano(ano_relatorio, 'HTM_02')

        # Criar o Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=relatorio_censo_{ano_relatorio}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = f"Censo {ano_relatorio}"

        # Estilos
        fonte_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        fonte_padrao = Font(name="Calibri", size=10)
        fonte_negrito = Font(name="Calibri", size=10, bold=True)
        alinhamento_central = Alignment(horizontal="center", vertical="center")
        fundo_titulo = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        fundo_dia = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")
        fundo_mes_claro = PatternFill(start_color="F2F8EE", end_color="F2F8EE", fill_type="solid")
        fundo_mes_escuro = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")
       

        borda_titulo = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        borda_dia = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        borda_mes = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style=None)
        )

        borda_qtd = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )

        borda_valor = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )

        borda_ext = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        borda_int = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        borda_total = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )



        # Aplicar bordas de dois em dois para colunas (B e C) até (Y e Z)
        for start_col in range(2, 26, 2):  # Colunas 2 (B) até 26 (Z), de 2 em 2
            for row in range(4, 35):  # Linhas de 4 a 34
                for col in range(start_col, start_col + 2):  # Colunas de par em par (ex: B e C, D e E)
                    if row == 4:  # Primeira linha com borda superior grossa
                        ws.cell(row=row, column=col).border = Border(
                            left=borda_ext.left if col == start_col else borda_int.left,
                            right=borda_int.right if col == start_col else borda_ext.right,
                            top=borda_ext.top,  # Superior grossa
                            bottom=borda_int.bottom  # Inferior fina
                        )
                    elif row == 34:  # Última linha com borda inferior grossa
                        ws.cell(row=row, column=col).border = Border(
                            left=borda_ext.left if col == start_col else borda_int.left,
                            right=borda_int.right if col == start_col else borda_ext.right,
                            top=borda_int.top,  # Superior fina
                            bottom=borda_ext.bottom  # Inferior grossa
                        )
                    else:  # Linhas intermediárias com bordas internas
                        ws.cell(row=row, column=col).border = Border(
                            left=borda_ext.left if col == start_col else borda_int.left,
                            right=borda_int.right if col == start_col else borda_ext.right,
                            top=borda_int.top,  # Superior fina
                            bottom=borda_int.bottom  # Inferior fina
                        )      


        # Título principal
        ws.merge_cells('A1:Y1')
        titulo_cell = ws['A1']
        titulo_cell.value = "HÓSPEDES E VALORES ARRECADADOS"
        titulo_cell.font = fonte_titulo
        titulo_cell.alignment = alinhamento_central
        titulo_cell.fill = fundo_titulo
        titulo_cell.border = borda_titulo

        # A1:Y1 corresponde a colunas 1 a 25
        for col in range(1, 26):  
            ws.cell(row=1, column=col).border = borda_titulo

        # Adicionar borda externa ao TOTAL
        ws.cell(row=35, column=1).border = borda_total    

        # Aplicar bordas de dois em dois na linha 35 (colunas B e C até Y e Z)
        for start_col in range(2, 26, 2):  # Colunas 2 (B) até 26 (Z), de 2 em 2
            for col in range(start_col, start_col + 2):  # Colunas de par em par (ex: B e C, D e E)
                ws.cell(row=35, column=col).border = Border(
                    left=borda_ext.left if col == start_col else borda_int.left,  # Esquerda grossa apenas na primeira coluna do par
                    right=borda_int.right if col == start_col else borda_ext.right,  # Direita fina na primeira coluna, grossa na última
                    top=borda_int.top,  # Superior fina
                    bottom=borda_ext.bottom  # Inferior grossa
                )


        # Cabeçalho DIA + Meses
        ws.merge_cells('A2:A3')
        ws['A2'] = "DIA"
        ws['A2'].font = fonte_negrito
        ws['A2'].alignment = alinhamento_central
        ws['A2'].fill = fundo_dia

        # Aplicar bordas ao intervalo A4:A34
        for row in range(4, 35):
            if row == 4:  # Primeira célula com borda superior grossa
                ws.cell(row=row, column=1).border = Border(
                    left=borda_ext.left,
                    right=borda_ext.right,
                    top=borda_ext.top,
                    bottom=borda_int.bottom
                )
            elif row == 34:  # Última célula com borda inferior grossa
                ws.cell(row=row, column=1).border = Border(
                    left=borda_ext.left,
                    right=borda_ext.right,
                    top=borda_int.top,
                    bottom=borda_ext.bottom
                )
            else:  # Células intermediárias com bordas internas
                ws.cell(row=row, column=1).border = Border(
                    left=borda_ext.left,
                    right=borda_ext.right,
                    top=borda_int.top,
                    bottom=borda_int.bottom
                )

        # Adicionar borda externa ao cabeçalho DIA
        for row in range(2, 4):
            ws.cell(row=row, column=1).border = borda_dia

        colunas = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
                   "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]

        col_idx = 2
        total_qtd = [0] * 12
        total_valor = [0.0] * 12

        for i, mes in enumerate(colunas):
            fundo_mes = fundo_mes_claro if i % 2 == 0 else fundo_mes_escuro

            # Borda do título do mês
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
            ws.cell(row=2, column=col_idx, value=mes).font = fonte_negrito
            ws.cell(row=2, column=col_idx).alignment = alinhamento_central
            ws.cell(row=2, column=col_idx).fill = fundo_mes
            ws.cell(row=2, column=col_idx).border = borda_mes
            ws.cell(row=2, column=col_idx + 1).border = borda_mes

            # Subtítulo QTD
            ws.cell(row=3, column=col_idx, value="QTD").font = fonte_negrito
            ws.cell(row=3, column=col_idx).alignment = alinhamento_central
            ws.cell(row=3, column=col_idx).fill = fundo_mes
            ws.cell(row=3, column=col_idx).border = borda_qtd

            # Subtítulo VALOR
            ws.cell(row=3, column=col_idx + 1, value="VALOR").font = fonte_negrito
            ws.cell(row=3, column=col_idx + 1).alignment = alinhamento_central
            ws.cell(row=3, column=col_idx + 1).fill = fundo_mes
            ws.cell(row=3, column=col_idx + 1).border = borda_valor

            col_idx += 2

        # Preencher os dados
        for dia in range(1, 32):
            ws.cell(row=dia + 3, column=1, value=dia).font = fonte_padrao
            ws.cell(row=dia + 3, column=1).fill = fundo_dia
            ws.cell(row=dia + 3, column=1).alignment = alinhamento_central

            col_idx = 2
            for mes in range(1, 13):
                if dia <= len(censo_htm_01[mes]) and dia <= len(censo_htm_02[mes]):
                    qtde = censo_htm_01[mes][dia - 1]['qtde'] + censo_htm_02[mes][dia - 1]['qtde']
                    valor = censo_htm_01[mes][dia - 1]['valor'] + censo_htm_02[mes][dia - 1]['valor']
                else:
                    qtde = 0
                    valor = 0.0

                # Preencher as células de QTDE e VALOR
                ws.cell(row=dia + 3, column=col_idx, value=qtde).alignment = alinhamento_central
                ws.cell(row=dia + 3, column=col_idx + 1, value=valor).number_format = '#,##0.00'
                ws.cell(row=dia + 3, column=col_idx).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro
                ws.cell(row=dia + 3, column=col_idx + 1).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro

                total_qtd[mes - 1] += qtde
                total_valor[mes - 1] += valor
                col_idx += 2

        # Linha TOTAL
        ws.cell(row=35, column=1, value="TOTAL").font = fonte_negrito
        ws.cell(row=35, column=1).alignment = alinhamento_central
        ws.cell(row=35, column=1).fill = fundo_dia

        col_idx = 2
        for mes in range(12):
            fundo_mes = fundo_mes_claro if mes % 2 == 0 else fundo_mes_escuro

            ws.cell(row=35, column=col_idx, value=total_qtd[mes]).font = fonte_negrito
            ws.cell(row=35, column=col_idx + 1, value=total_valor[mes]).font = fonte_negrito
            ws.cell(row=35, column=col_idx + 1).number_format = '#,##0.00'
            ws.cell(row=35, column=col_idx).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx + 1).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx).fill = fundo_mes
            ws.cell(row=35, column=col_idx + 1).fill = fundo_mes

            col_idx += 2

        wb.save(response)
        return response












@login_required
def relatorio_censo_ano(request):
    return render(request, 'portal/relatorio_censo_ano.html')









###############################################################################


from datetime import datetime, timedelta
from collections import OrderedDict
from calendar import monthrange
from django.db.models import Q

def obter_agenda(mes, ano, hotel):
    ocupacao_hotel = OrderedDict()
    dias_no_mes = monthrange(ano, mes)[1]

    # Inicializar todos os dias do mês para cada quarto
    for dia in range(1, dias_no_mes + 1):
        ocupacao_hotel[dia] = {}
        for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
            ocupacao_hotel[dia][str(quarto)] = {
                'ocupado': False,
                'media_diaria': ""
            }

    # Consultar todas as reservas, excluindo status "Pendente" ou "Recusada"
    reservas = BaseDados.objects.filter(
        mhex=hotel,
        entrada__year=ano,
        entrada__month=mes,
        status_reserva__in=['Confirmada', 'Pago']  # Certifique-se de que esses são os status corretos
    )

    for reserva in reservas:
        diarias = reserva.diarias if reserva.diarias > 0 else 0
        quarto = str(reserva.uh)
        valor_total = reserva.valor_total
        media_diaria = valor_total / diarias if diarias > 0 else 0

        # Calcular e marcar os dias da reserva como ocupados
        data_entrada = reserva.entrada
        data_saida = data_entrada + timedelta(days=diarias)

        while data_entrada < data_saida:
            dia_reserva = data_entrada.day
            if dia_reserva <= dias_no_mes:
                if quarto in ocupacao_hotel[dia_reserva]:
                    ocupacao_hotel[dia_reserva][quarto]['ocupado'] = True
                    ocupacao_hotel[dia_reserva][quarto]['media_diaria'] = "{:.2f}".format(media_diaria)
            data_entrada += timedelta(days=1)

    # Ordenar o dicionário por dia
    ocupacao_hotel = OrderedDict(sorted(ocupacao_hotel.items()))

    return ocupacao_hotel




class Visualizar_Agenda_Ocupacao(View):
    def get(self, request, *args, **kwargs):
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')

        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes e ano são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        ocupacao_htm_01 = obter_agenda(mes_relatorio, ano_relatorio, 'HTM_01')
        ocupacao_htm_02 = obter_agenda(mes_relatorio, ano_relatorio, 'HTM_02')

        return render(request, 'portal/ocupacao.html', {
            'ocupacao_htm_01': ocupacao_htm_01,
            'ocupacao_htm_02': ocupacao_htm_02,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,
        })

@login_required
def agenda_ocupacao(request):    
    return render(request, 'portal/relatorio_agenda.html')  


@login_required
def relatorio_todos(request):    
    return render(request, 'portal/relatorio_todos.html') 


@login_required
def consultar_todos(request):
    if request.method == 'GET':
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')     

        # Armazenar os parâmetros na sessão
        request.session['mes_relatorio'] = mes_relatorio
        request.session['ano_relatorio'] = ano_relatorio   
        
        # Verificar se mes_relatorio e ano_relatorio não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio e ano_relatorio são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")
        
        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")
        
        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(            
            saida__month=mes_relatorio,
            saida__year=ano_relatorio,
            mhex="HTM_01",
        ).exclude(status_reserva="Recusada").order_by('entrada')      

          

        consultar_htm2 = BaseDados.objects.filter(            
            saida__month=mes_relatorio,
            saida__year=ano_relatorio,
            mhex="HTM_02",
        ).exclude(status_reserva="Recusada").order_by('entrada')

        

        # Atualização em lote dos objetos BaseDados
        if consultar_htm1.exists():
            consultar_htm1.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        if consultar_htm2.exists():
            consultar_htm2.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        context = {
            'consultar_htm1': consultar_htm1,
            'consultar_htm2': consultar_htm2,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,           

        }
        return render(request, 'portal/consultar_todos.html', context)
    else:
        return HttpResponseBadRequest("Método de requisição inválido.")

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import BaseDados
from .forms import ReservasForm

@login_required
def editar_todos(request, reservas_pk):
    editar = get_object_or_404(BaseDados, pk=reservas_pk)
    mes_relatorio = request.session.get('mes_relatorio')
    ano_relatorio = request.session.get('ano_relatorio')

    if request.method == 'POST':
        post_data = request.POST.copy()

        # Mantém os valores existentes no banco, alterando apenas os campos permitidos
        if "entrada" in post_data and post_data["entrada"]:
            editar.entrada = post_data["entrada"]
        if "saida" in post_data and post_data["saida"]:
            editar.saida = post_data["saida"]
        if "status_reserva" in post_data and post_data["status_reserva"]:
            editar.status_reserva = post_data["status_reserva"]
        if "mhex" in post_data and post_data["mhex"]:
            editar.mhex = post_data["mhex"]
        if "uh" in post_data and post_data["uh"]:
            editar.uh = post_data["uh"]
        if "valor_ajuste" in post_data and post_data["valor_ajuste"]:
            editar.valor_ajuste = post_data["valor_ajuste"].replace(',', '.')

        # Salva apenas os campos editáveis
        editar.save()

        # Redireciona para consultar_todos mantendo os filtros de data
        redirect_url = reverse('consultar_todos') + f'?mes_relatorio={mes_relatorio}&ano_relatorio={ano_relatorio}'
        return redirect(redirect_url)

    # Carrega os dados do banco corretamente
    form = ReservasForm(instance=editar)

    context = {
        'form': form,
        'reserva': editar,
        'mes_relatorio': mes_relatorio,
        'ano_relatorio': ano_relatorio,
    }
    return render(request, 'portal/editar_todos.html', context)


# @login_required
# def editar_todos(request, reservas_pk):
#     # Aqui usamos get_object_or_404 para garantir que se o objeto não for encontrado,
#     # uma página de erro 404 será retornada.
#     editar = get_object_or_404(BaseDados, pk=reservas_pk)
#     mes_relatorio = request.session.get('mes_relatorio')
#     ano_relatorio = request.session.get('ano_relatorio')
    
    
#     # Se estamos lidando com uma requisição POST, significa que o formulário foi submetido
#     if request.method == 'POST':
#         # Ajuste do valor_ajuste
#         if 'valor_ajuste' in request.POST:
#             valor_ajuste = request.POST['valor_ajuste'].replace(',', '.')
#             request.POST = request.POST.copy()
#             request.POST['valor_ajuste'] = valor_ajuste
        
#         # Passamos a instância da reserva que queremos editar para o formulário,
#         # junto com os dados submetidos (request.POST)
#         form = ReservasForm(request.POST, instance=editar)
#         print("Dados recebidos do formulário via POST:")
#         print(request.POST)
#         if 'status_reserva' in request.POST:
#             status_reserva = request.POST['status_reserva']
#             print("Status da reserva recebido:", status_reserva)
#             # Atribuir o valor do campo status_reserva à instância do formulário
#             form.instance.status_reserva = status_reserva
        
#         # Verificamos se o formulário é válido
#         if form.is_valid():
#             print("Formulário válido. Tentando salvar...")
#             # Se for válido, salvamos as alterações feitas na reserva
#             form.save()
#             print("Reserva salva com sucesso!")
#             # Construir a URL de redirecionamento para consultar_todos com os parâmetros de consulta
#             redirect_url = reverse('consultar_todos')
#             redirect_url += f'?mes_relatorio={mes_relatorio}&ano_relatorio={ano_relatorio}'
#             # Redirecionar para consultar_todos com os parâmetros de consulta
#             return redirect(redirect_url)
#         else:
#             print("Formulário inválido. Erros de validação:")
#             print(form.errors)
#     else:
#         # Se a requisição for um GET, criamos um formulário preenchido com as informações
#         # da reserva que queremos editar
#         form = ReservasForm(instance=editar)
    
#     # Passamos o formulário preenchido para o template
#     context = {
#         'form': form, 
#         'reserva': editar,
#         'mes_relatorio': mes_relatorio,
#         'ano_relatorio': ano_relatorio, 
#     }
#     return render(request, 'portal/editar_todos.html', context)



from collections import defaultdict
from datetime import datetime
from calendar import monthrange

from collections import defaultdict
from calendar import monthrange
from datetime import date
from .models import BaseDados  # Ajuste conforme o caminho correto do seu modelo

def obter_relatorio_hospedes_valor(mes, ano, hotel):
    relatorio = defaultdict(lambda: {'total_hospedes': 0, 'valor_total': Decimal('0.0')})
    dias_no_mes = monthrange(ano, mes)[1]

    print(f"Verificando hotel: {hotel}, mês: {mes}, ano: {ano}")

    for dia in range(1, dias_no_mes + 1):
        data_atual = date(ano, mes, dia)

        reservas_ativas_no_dia = BaseDados.objects.filter(
            entrada__lte=data_atual,
            saida__gte=data_atual,
            mhex=hotel
        )

        print(f"Dia {dia}: encontradas {reservas_ativas_no_dia.count()} reservas ativas")

        for reserva in reservas_ativas_no_dia:
            qtde_hosp = Decimal(reserva.qtde_hosp if reserva.qtde_hosp > 0 else 1)
            valor_por_pessoa = Decimal(reserva.valor_total) / qtde_hosp

            relatorio[dia]['total_hospedes'] += qtde_hosp
            relatorio[dia]['valor_total'] += valor_por_pessoa * qtde_hosp

            print(f"Reserva ID {reserva.id}: qtde_hosp={qtde_hosp}, valor_total={reserva.valor_total}")

    # Preparação para visualização dos resultados acumulados
    for dia, dados in relatorio.items():
        print(f"Dia {dia}: total_hospedes={dados['total_hospedes']}, valor_total={dados['valor_total']}")

    relatorio_formatado = {
        dia: {
            'total_hospedes': relatorio[dia]['total_hospedes'],
            'valor_total': "{:.2f}".format(relatorio[dia]['valor_total'])
        } for dia in relatorio
    }

    return relatorio_formatado  


def relatorio_anual_hoteis(request):
    ano = 2024
    hoteis = ['HTM_01', 'HTM_02']
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    relatorios = []

    # Criar uma lista de dias para cada mês (assumindo que todos os meses têm 31 dias)
    dias_no_mes = list(range(1, 32))

    for hotel in hoteis:
        dados_hotel = {'nome': hotel, 'dados_meses': [], 'dias_no_mes': dias_no_mes}
        for mes_idx, mes in enumerate(meses, start=1):
            # Usando a função correta para recuperar os dados
            dados_mes = obter_relatorio_hospedes_valor(mes_idx, ano, hotel)
            dados_hotel['dados_meses'].append({
                'mes': mes,
                'total_hospedes': dados_mes.get('total_hospedes', {}),
                'valor_total': dados_mes.get('valor_total', {})
            })
        relatorios.append(dados_hotel)

    context = {
        'relatorios': relatorios,
        'ano': ano,
    }
    return render(request, 'portal/consultar_censo1.html', context)


########## RELATORIO HOSPEDAGEM DIARIA POR NOMES

from django.views import View
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import timedelta
from .models import BaseDados  # Importa o modelo correto




class RelatorioHospedagemView(View):
    def get(self, request, *args, **kwargs):
        # Buscar reservas com status_reserva="Pago" e ordenar por entrada
        reservas = BaseDados.objects.filter(status_reserva="Pago").order_by('entrada')

        # Log para verificar os dados ordenados
        print("Verificando datas ordenadas:")
        for reserva in reservas:
            print(f"ID: {reserva.id}, Entrada: {reserva.entrada}, Saída: {reserva.saida}")

        # Criar arquivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Hospedagem"

        # Estilos
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Cabeçalho
        headers = ["ID", "Data", "Nome do Hóspede", "Valor Diário", "Diárias", "Valor Total"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            ws.column_dimensions[cell.column_letter].width = 20

        # Preencher os dados
        row_num = 2
        for reserva in reservas:
            # Garantir que as datas de entrada e saída são válidas
            if reserva.entrada and reserva.saida and reserva.valor_total:
                dias_hospedagem = (reserva.saida - reserva.entrada).days
                if dias_hospedagem <= 0:
                    continue  # Ignorar reservas com datas inválidas

                # Calcular valor diário
                valor_diario = reserva.valor_total / dias_hospedagem

                # Iterar pelos dias de hospedagem (sem contar o dia da saída)
                for i in range(dias_hospedagem):
                    data_hospedagem = reserva.entrada + timedelta(days=i)
                    ws.cell(row=row_num, column=1, value=reserva.id)  # ID da reserva
                    ws.cell(row=row_num, column=2, value=data_hospedagem.strftime("%d/%m/%Y"))  # Data
                    ws.cell(row=row_num, column=3, value=reserva.nome)  # Nome do hóspede
                    ws.cell(row=row_num, column=4, value=round(valor_diario, 2))  # Valor diário
                    ws.cell(row=row_num, column=5, value=dias_hospedagem)  # Diárias
                    ws.cell(row=row_num, column=6, value=reserva.valor_total)  # Valor total
                    row_num += 1

        # Retornar como resposta HTTP
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="relatorio_hospedagem.xlsx"'
        wb.save(response)
        return response

############# RELATORIO CENSO POR UH OCUPADA

import datetime
from calendar import monthrange
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import BaseDados

# Função que processa o relatório
def obter_relatorio_uh(ano, hotel, mes_selecionado=None):
    from datetime import date, timedelta
    from calendar import monthrange

    # Inicializar o relatório
    relatorio = {mes: [{'dia': dia, 'qtde': 0, 'uh': 0} for dia in range(1, monthrange(ano, mes)[1] + 1)] for mes in range(1, 13)}

    # Buscar reservas válidas no banco de dados
    reservas = BaseDados.objects.filter(
        mhex=hotel,
        status_reserva="Pago",
        entrada__lte=date(ano, 12, 31),
        saida__gte=date(ano, 1, 1)
    )

    # Processar cada reserva
    for reserva in reservas:
        if not reserva.entrada or not reserva.saida or not reserva.qtde_quartos:
            # Ignorar reservas inválidas
            continue

        # Garantir que a reserva está dentro do ano solicitado
        data_inicio = max(reserva.entrada, date(ano, 1, 1))  # O início da reserva dentro do ano
        data_fim = min(reserva.saida, date(ano, 12, 31))  # O fim da reserva dentro do ano

        if data_inicio > data_fim:
            # Ignorar reservas fora do ano
            continue

        # Iterar apenas pelos dias da reserva dentro do ano
        data_atual = data_inicio
        while data_atual < data_fim:
            mes = data_atual.month
            dia = data_atual.day

            # Somar os valores no relatório
            relatorio[mes][dia - 1]['qtde'] += reserva.qtde_hosp or 0
            relatorio[mes][dia - 1]['uh'] += reserva.qtde_quartos or 0

            data_atual += timedelta(days=1)

    # Zerar dias sem ocupação
    for mes in relatorio:
        for dia in relatorio[mes]:
            if dia['uh'] == 0:
                dia['qtde'] = 0  # Zerar hóspedes para dias sem quartos ocupados

    # Retornar o relatório completo
    return relatorio


class Visualizar_Censo_UH(View):
    def get(self, request, *args, **kwargs):
        ano_relatorio = request.GET.get('ano_relatorio')

        if not ano_relatorio:
            return render(request, 'portal/relatorio_censo_ano.html', {
                'error_message': "Por favor, selecione o ano para gerar o relatório."
            })

        try:
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Ano deve ser um número inteiro.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Gerar relatórios para os hotéis HTM_01 e HTM_02
        censo_htm_01 = obter_relatorio_uh(ano_relatorio, 'HTM_01')  # Aqui foi corrigido
        censo_htm_02 = obter_relatorio_uh(ano_relatorio, 'HTM_02')  # Aqui foi corrigido

        # Criar o Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=relatorio_censo_uh_{ano_relatorio}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = f"Censo {ano_relatorio}"

        # Estilos
        fonte_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        fonte_padrao = Font(name="Calibri", size=10)
        fonte_negrito = Font(name="Calibri", size=10, bold=True)
        alinhamento_central = Alignment(horizontal="center", vertical="center")
        fundo_titulo = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        fundo_dia = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")
        fundo_mes_claro = PatternFill(start_color="F2F8EE", end_color="F2F8EE", fill_type="solid")
        fundo_mes_escuro = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")

        borda_titulo = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        borda_mes = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style=None)
        )

        borda_qtd = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )

        borda_uh = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )

        # Título principal
        ws.merge_cells('A1:Y1')
        titulo_cell = ws['A1']
        titulo_cell.value = "UNIDADES HABITACIONAIS"
        titulo_cell.font = fonte_titulo
        titulo_cell.alignment = alinhamento_central
        titulo_cell.fill = fundo_titulo
        titulo_cell.border = borda_titulo

        # Cabeçalho DIA + Meses
        ws.merge_cells('A2:A3')
        ws['A2'] = "DIA"
        ws['A2'].font = fonte_negrito
        ws['A2'].alignment = alinhamento_central
        ws['A2'].fill = fundo_dia

        colunas = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
                   "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]

        col_idx = 2
        total_qtd = [0] * 12
        total_uh = [0] * 12

        for i, mes in enumerate(colunas):
            fundo_mes = fundo_mes_claro if i % 2 == 0 else fundo_mes_escuro

            # Borda do título do mês
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
            ws.cell(row=2, column=col_idx, value=mes).font = fonte_negrito
            ws.cell(row=2, column=col_idx).alignment = alinhamento_central
            ws.cell(row=2, column=col_idx).fill = fundo_mes
            ws.cell(row=2, column=col_idx).border = borda_mes
            ws.cell(row=2, column=col_idx + 1).border = borda_mes

            # Subtítulo UH DISPONIVEL
            ws.cell(row=3, column=col_idx, value="TOTAL DE UH DO DIA").font = fonte_negrito
            ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=3, column=col_idx).fill = fundo_mes
            ws.cell(row=3, column=col_idx).border = borda_qtd

            # Subtítulo UH OCUPADA
            ws.cell(row=3, column=col_idx + 1, value="UH OCUPADA NO DIA").font = fonte_negrito
            ws.cell(row=3, column=col_idx + 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=3, column=col_idx + 1).fill = fundo_mes
            ws.cell(row=3, column=col_idx + 1).border = borda_uh

            col_idx += 2

        # Preencher os dados
        for dia in range(1, 32):
            ws.cell(row=dia + 3, column=1, value=dia).font = fonte_padrao
            ws.cell(row=dia + 3, column=1).fill = fundo_dia
            ws.cell(row=dia + 3, column=1).alignment = alinhamento_central

            col_idx = 2
            for mes in range(1, 13):
                uh = sum(
                    censo[mes][dia - 1]['uh']
                    for censo in [censo_htm_01, censo_htm_02]
                    if mes in censo and dia <= len(censo[mes])
                )

                # Deixar QTD em branco e manter bordas e estilos
                ws.cell(row=dia + 3, column=col_idx, value="").alignment = alinhamento_central
                ws.cell(row=dia + 3, column=col_idx).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro
                ws.cell(row=dia + 3, column=col_idx).border = borda_qtd

                # Preencher UH
                ws.cell(row=dia + 3, column=col_idx + 1, value=uh).alignment = alinhamento_central
                ws.cell(row=dia + 3, column=col_idx + 1).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro
                ws.cell(row=dia + 3, column=col_idx + 1).border = borda_uh

                total_uh[mes - 1] += uh
                col_idx += 2

        # Linha TOTAL
        ws.cell(row=35, column=1, value="TOTAL").font = fonte_negrito
        ws.cell(row=35, column=1).alignment = alinhamento_central
        ws.cell(row=35, column=1).fill = fundo_dia

        col_idx = 2
        for mes in range(12):
            fundo_mes = fundo_mes_claro if mes % 2 == 0 else fundo_mes_escuro

            ws.cell(row=35, column=col_idx, value=total_qtd[mes]).font = fonte_negrito
            ws.cell(row=35, column=col_idx + 1, value=total_uh[mes]).font = fonte_negrito
            ws.cell(row=35, column=col_idx).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx + 1).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx).fill = fundo_mes
            ws.cell(row=35, column=col_idx + 1).fill = fundo_mes

            col_idx += 2

        wb.save(response)
        return response







