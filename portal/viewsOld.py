from datetime import datetime
from io import BytesIO
from unittest import case
from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django import forms
from django.core.paginator import Paginator
from django.views import View
from portal.models import BaseDados, Produto, Precos_graduacao_vinculo, Precos_status_graduacao
from .forms import ReservasForm
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseNotAllowed, JsonResponse
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect
import logging
from decimal import Decimal
from django.template.defaulttags import register
from django.db.models import Sum
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import locale
from datetime import datetime
from django.db.models import F
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
import xlsxwriter
from django.http import Http404
from django.db.models import Case, When, Value, F, IntegerField, Q
from django.views.generic import View
import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from datetime import timedelta
from django.shortcuts import render
from django.utils import timezone
from calendar import monthrange
from django.db.models import Avg
from datetime import datetime, timedelta, date
import decimal
from django.db import transaction
from django.http import JsonResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt




@register.filter
def get_range(value):
    return range(value)


@login_required
def consultar_reservas(request):
    consultar = BaseDados.objects.filter(status_reserva="Pendente")  # Filtra por status "pendente"
    print(consultar)
    context = {
        'consultar': consultar,
    }
    return render(request, 'portal/reservas.html', context)

# def reserva_externa(request):
#     if request.method == 'POST':
#         form = ReservasForm(request.POST)
#         if form.is_valid():
#             obj = form.save()
#             print("Objeto salvo:", obj)
#             form.save()
#             # return redirect('consultar_reservas')
#             return redirect('reserva_externa')
#         else:
#             # Se o formulário for inválido, cairemos aqui e os erros serão impressos.
#             print("Formulário inválido", form.errors)
#     else:
#         # Se a requisição não for POST, um formulário em branco será fornecido.
#         form = ReservasForm()

#     # O contexto é passado para o template tanto se o formulário for válido (e redirecionado antes),
#     # quanto se não for POST ou se for inválido.
#     context = {'form': form}
#     return render(request, 'portal/index.html', context)


# @transaction.atomic
# def reserva_externa(request):
#     if request.method == 'POST':
#         form = ReservasForm(request.POST)
#         if form.is_valid():
#             obj = form.save()
#             print("Objeto salvo:", obj, "ID:", obj.id)
#             return redirect('reserva_externa')
#         else:
#             print("Formulário inválido", form.errors)
#     else:
#         form = ReservasForm()

#     context = {'form': form}
#     return render(request, 'portal/index.html', context)

@transaction.atomic
def reserva_externa(request):
    if request.method == 'POST':
        form = ReservasForm(request.POST)
        if form.is_valid():
            try:
                obj = form.save()  # Salva o formulário e obtém o objeto salvo
                print("Objeto salvo:", obj, "ID:", obj.id)
                return JsonResponse({"success": True, "message": "Cadastro realizado com sucesso!"})
            except Exception as e:
                # Em caso de erro no salvamento, retorna uma resposta de falha
                print("Erro ao salvar:", str(e))
                return JsonResponse({"success": False, "message": "Erro ao salvar os dados!"}, status=500)
        else:
            # Se o formulário for inválido, retorna os erros em formato JSON
            errors = form.errors.as_json()
            print("Formulário inválido", errors)
            return JsonResponse({"success": False, "message": "Dados do formulário inválidos!", "errors": errors}, status=400)
    else:
        # Se a requisição não for POST, retorna um HTML normal
        form = ReservasForm()
        context = {'form': form}
        return render(request, 'portal/index.html', context)



@login_required
def fazer_reservas(request):
    if request.method == 'POST':
        form = ReservasForm(request.POST)
        if form.is_valid():
            obj = form.save()
            # print("Objeto salvo:", obj)
            # form.save()
            # return redirect('consultar_reservas')
            return redirect('consultar_reservas')
        else:
            # Se o formulário for inválido, cairemos aqui e os erros serão impressos.
            print("Formulário inválido", form.errors)
    else:
        # Se a requisição não for POST, um formulário em branco será fornecido.
        form = ReservasForm()

    # O contexto é passado para o template tanto se o formulário for válido (e redirecionado antes),
    # quanto se não for POST ou se for inválido.
    context = {'form': form}
    return render(request, 'portal/reserva_add.html', context)

@login_required
def editar_reservas(request, reservas_pk):
    editar = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        form = ReservasForm(request.POST, instance=editar)
        print("Dados recebidos do formulário:", request.POST)  # Adiciona mensagem de log para verificar os dados do formulário
        if 'status_reserva' in request.POST:
            status_reserva = request.POST['status_reserva']
            print("Status da reserva recebido:", status_reserva)  # Adiciona mensagem de log para verificar o status da reserva
            form.instance.status_reserva = status_reserva
        
        if form.is_valid():
            form.save()
            print("Formulário válido. Salvando alterações...")  # Adiciona mensagem de log para verificar se o formulário é válido
            return redirect('consultar_reservas')
        else:
            print("Formulário inválido. Erros:", form.errors)  # Adiciona mensagem de log para verificar os erros de validação do formulário
            # Se o formulário for inválido, podemos retornar uma resposta HTTP com os erros para ajudar na depuração
            return HttpResponse("Formulário inválido. Erros: {}".format(form.errors), status=400)
    else:
        form = ReservasForm(instance=editar)
    
    context = {'form': form, 'reserva': editar}
    return render(request, 'portal/reserva_edit.html', context)

@login_required
def recepcao_checkin(request):
    consultar = BaseDados.objects.filter(status_reserva="Aprovada")  # Filtra por status "pendente"
    consultar_checkin = BaseDados.objects.filter(status_reserva="Checkin")
    # print(consultar)    
    context = {
        'consultar': consultar,
        'consultar_checkin': consultar_checkin,
    }
    return render(request, 'portal/recepcao_checkin.html', context)  

@login_required
def recepcao_checkout(request):
    consultar = BaseDados.objects.filter(status_reserva="Aprovada")  # Filtra por status "pendente"
    consultar_checkin = BaseDados.objects.filter(status_reserva="Checkin")
    # print(consultar)    
    context = {
        'consultar': consultar,
        'consultar_checkin': consultar_checkin,
    }
    return render(request, 'portal/recepcao_checkout.html', context)      



@csrf_protect
@login_required
def editar_checkin(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        novo_status = request.POST.get('status_reserva')
        reserva.status_reserva = novo_status
        
        # Calcula as diárias
        if reserva.entrada and reserva.saida:
            try:
                formato_data = "%d/%m/%Y"
                data_entrada_str = reserva.entrada.strftime(formato_data)  # Convertendo para string
                data_saida_str = reserva.saida.strftime(formato_data)  # Convertendo para string

                data_entrada = datetime.strptime(data_entrada_str, formato_data)
                data_saida = datetime.strptime(data_saida_str, formato_data)
                
                diarias = (data_saida - data_entrada).days
                reserva.diarias = diarias
                print(data_entrada)
                print(data_saida)
                print(reserva.diarias)
            except ValueError as e:
                print(f"Erro ao calcular diárias: {e}")
                reserva.diarias = 0
        
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao_checkin'))
    
    return render(request, 'portal/checkin.html', {'reserva': reserva})



def editar_checkout(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        novo_status = request.POST.get('status_reserva')
        novo_formapg = request.POST.get('forma_pagamento')
        reserva.status_reserva = novo_status
        reserva.forma_pagamento = novo_formapg
        
        # Verifica se o checkbox foi marcado
        if 'pagante_checkbox' in request.POST:
            print("Checkbox foi marcado")
            nome_pagante = request.POST.get('nome_pagante')
            cpf_pagante = request.POST.get('cpf_pagante')
            print("Nome do pagante:", nome_pagante)
            print("CPF do pagante:", cpf_pagante)
        else:
            # Se o checkbox não foi marcado, usa os dados existentes
            print("Checkbox não foi marcado")
            nome_pagante = reserva.nome
            cpf_pagante = reserva.cpf
        
        # Atualiza os campos na reserva
        reserva.nome_pagante = nome_pagante
        reserva.cpf_pagante = cpf_pagante
        
        # Salva a reserva
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao_checkout'))
    
    return render(request, 'portal/checkout.html', {'reserva': reserva})



@csrf_protect

def editar_consumacao(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)

    print("Valores do objeto reserva antes de renderizar o formulário:", reserva.qtde_agua, reserva.qtde_refri, reserva.qtde_cerveja, reserva.status, reserva.graduacao, reserva.valor_hosp)
    
    if request.method == 'POST':
        novo_agua = request.POST.get('qtde_agua')
        novo_refri = request.POST.get('qtde_refri')
        novo_cerveja = request.POST.get('qtde_cerveja')
        novo_pet = request.POST.get('qtde_pet')

        status = reserva.status
        graduacao = reserva.graduacao

        # Consulta os preços com base no status e na graduação
        try:
            if status == 'CIVIL':
                preco_status_graduacao = Precos_status_graduacao.objects.get(status=status, graduacao=graduacao)
                reserva.valor_hosp = preco_status_graduacao.valor
            elif status in ['MILITAR DA ATIVA', 'MILITAR DA RESERVA', 'DEP ACOMPANHADO', 'DEP DESACOMPANHADO', 'PENSIONISTA']:                
                preco_status_graduacao = Precos_status_graduacao.objects.get(status=status, graduacao=graduacao)
                reserva.valor_hosp = preco_status_graduacao.valor
            else:
                # Se não encontrar um preço correspondente, define valor_hosp como zero
                reserva.valor_hosp = 0  # Ou algum outro valor padrão que faça sentido em seu contexto
        except Precos_status_graduacao.DoesNotExist:
            # Se o objeto não for encontrado, define valor_hosp como zero
            reserva.valor_hosp = 0
        
        # Exemplo de cálculo para um acompanhante
        vinculo_acomps = [reserva.vinculo_acomp1, reserva.vinculo_acomp2, reserva.vinculo_acomp3, reserva.vinculo_acomp4, reserva.vinculo_acomp5]
        for i, vinculo_acomp in enumerate(vinculo_acomps):
            if vinculo_acomp:
                try:
                    # Consulta o preço do vínculo e graduação na tabela Precos_graduacao_vinculo
                    preco_vinculo = Precos_graduacao_vinculo.objects.get(graduacao=graduacao, vinculo=vinculo_acomp)
                    setattr(reserva, f'valor_acomp{i+1}', preco_vinculo.valor)
                except Precos_graduacao_vinculo.DoesNotExist:
                    setattr(reserva, f'valor_acomp{i+1}', 0)
        # Cálculo dos novos valores conforme solicitado
        if isinstance(reserva.valor_hosp, (int, float, Decimal)):
            reserva.valor_dia = (reserva.valor_hosp + reserva.valor_acomp1 + reserva.valor_acomp2 + reserva.valor_acomp3 + reserva.valor_acomp4 + reserva.valor_acomp5)
        else:
            # Se reserva.valor_hosp não for numérico, defina reserva.valor_dia como zero ou algum outro valor padrão
            reserva.valor_dia = 0
                   

        # print("Tipo de dado do valor:", type(preco_status_graduacao.valor))
        print("Valor do hospede:", reserva.valor_hosp)
        print("Valores recebidos do formulário:")
        print("Água:", novo_agua)
        print("Refri:", novo_refri)
        print("Cerveja:", novo_cerveja)
        reserva.qtde_acomp = reserva.qtde_hosp - 1
        
        # Verifica se os valores não são None antes de converter para Decimal
        if novo_agua is not None:
            reserva.qtde_agua = Decimal(novo_agua)
        if novo_refri is not None:
            reserva.qtde_refri = Decimal(novo_refri)
        if novo_cerveja is not None:
            reserva.qtde_cerveja = Decimal(novo_cerveja)
        if novo_pet is not None:
            reserva.qtde_pet = Decimal(novo_pet)    

        
            
        # Busca os valores dos produtos no banco de dados
        produto_agua = Produto.objects.get(nome='Água')
        produto_refri = Produto.objects.get(nome='Refrigerante')
        produto_cerveja = Produto.objects.get(nome='Cerveja')   
        produto_pet = Produto.objects.get(nome='Pet')  
        
        if reserva.motivo_viagem == "Saúde":
            reserva.desc_saude = reserva.valor_dia / 2
            reserva.subtotal = reserva.desc_saude * reserva.diarias          
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao  
            reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
            reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
            reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
            if reserva.qtde_pet != 0:
                reserva.total_pet = reserva.diarias * produto_pet.valor if reserva.diarias is not None else Decimal(0)
            else:
                reserva.total_pet = 0;    
            reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja 
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao + reserva.total_pet
        else:
            reserva.desc_saude = 0
            reserva.subtotal = reserva.valor_dia * reserva.diarias          
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao  
            reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
            reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
            reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
            if reserva.qtde_pet != 0:
                reserva.total_pet = reserva.diarias * produto_pet.valor if reserva.diarias is not None else Decimal(0)
            else:
                reserva.total_pet = 0; 
            reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao + reserva.total_pet  


        # Calcula consumacao
        reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
        reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
        reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
        reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
        reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao + reserva.total_pet
        
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao_checkout'))
    
    return render(request, 'portal/consumacao.html', {'reserva': reserva})




@login_required
def relatorio_mensal(request):
    if request.method == 'GET':
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        print("Mês Relatório:", mes_relatorio)  # Verifique se os valores estão corretos
        print("Ano Relatório:", ano_relatorio)
        print("Forma Pagamento:", forma_pagamento)
        
        # Verificar se mes_relatorio e ano_relatorio não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio e ano_relatorio são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")
        
        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")
        
        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        )

        total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma'] 

          

        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        )

        total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma'] 

        # Atualização em lote dos objetos BaseDados
        if consultar_htm1.exists():
            consultar_htm1.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        if consultar_htm2.exists():
            consultar_htm2.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        context = {
            'consultar_htm1': consultar_htm1,
            'consultar_htm2': consultar_htm2,
            'total_htm1': total_htm1,
            'total_htm2': total_htm2,

        }
        return render(request, 'portal/relatorio_mensal.html', context)
    else:
        return HttpResponseBadRequest("Método de requisição inválido.")

@login_required
def relatorio_pagamento(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento.html', context) 

@login_required
def relatorio_pagamento_pix(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_pix.html', context) 

@login_required
def relatorio_pagamento_dinheiro(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_dinheiro.html', context) 





@login_required
def relatorio_pagamento_excel(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_excel.html', context) 



class RelatorioPix(View):
    nome_colunas_excel = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    cabecalho_fixo = 9
    cabecalho_dados1 = 1
    cabecalho_dados2 = 1
    linhas_dados1 = None
    linhas_dados2 = None
    mes_texto = None
    total_linhas = None
    total_htm1 = 0
    total_htm2 = 0
    dados1 = None
    dados2 = None
    mes_relatorio = None
    ano_relatorio = None


    def get(self, request, *args, **kwargs):
        self.mes_relatorio = request.GET.get('mes_relatorio')
        self.ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
        # print("Ano Relatório: auditoria", ano_relatorio)
        # print("Forma Pagamento: auditoria", forma_pagamento)

        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        
        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

        # Validação dos dados de entrada
        try:
            self.mes_relatorio = int(self.mes_relatorio)
            self.ano_relatorio = int(self.ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if self.mes_relatorio < 1 or self.mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
        print(self.mes_texto)

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        # Consulta para o segundo tipo de dados
        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        if not consultar_htm1.exists() and not consultar_htm2.exists():
            return render(request, 'portal/relatorio_pagamento_pix.html', {'exibir_modal': True})

        self.total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']  

        self.total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']

        # Combine os resultados das consultas em um único conjunto de dados
        self.dados1 = list(consultar_htm1)
        self.dados2 = list(consultar_htm2)

        self.linhas_dados1 = len(self.dados1) 
        self.linhas_dados2 = len(self.dados2)       
        self.total_linhas = self.linhas_dados1 + self.linhas_dados2

        # Gere a planilha
        workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

        # Configura o Excel
        self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

        # Configura o Estilo
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

        # Retorna a resposta HTTP com a planilha anexada
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

        # Salva o workbook como uma resposta HTTP
        workbook.save(response)
        return response

    def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'{ self.mes_texto}'

        # INCLUIR CABEÇALHO DADOS 1
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=10, column=col).value = coluna

        # INCLUIR DADOS 1
        self.escrever_dados(sheet, 11, self.dados1)

        # INCLUIR CABEÇALHO DADOS 2
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

        # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
        self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

        return workbook

    def escrever_dados(self, sheet, start_row, dados):
        # Escrever os dados
        for row, registro in enumerate(dados, start=start_row):
            for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
                valor = getattr(registro, coluna, None)
                if coluna == 'valor_total' and valor is not None:
                    try:
                        valor = float(valor)
                    except ValueError:
                        valor = None
                elif coluna == 'graduacao':
                # Verifica se o status é diferente de "MILITAR DA ATIVA" ou "MILITAR DA RESERVA"
                    if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
                        valor = "CIVIL"        
                sheet.cell(row=row, column=col).value = valor

    def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
        # Acessa a planilha ativa
        sheet = workbook.active

        # Larguras das colunas desejadas
        larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras_colunas.items():
            sheet.column_dimensions[col].width = width

        # Formatar coluna 'A' (Saida) para dd/mm/aaaa
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'

        # Aplicar estilo à coluna 'A' (Saida)
        for cell in sheet['A']:
            cell.style = date_format

        # Ocultar as linhas de grade
        sheet.sheet_view.showGridLines = False

        # Formatar coluna 'H' (valor_total) para ##.###,##
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

        # Aplicar estilo à coluna 'H' (valor_total)
        for cell in sheet['H']:
            cell.style = valor_format

        # Consulta para os dados do banco
        consultar_dados = self.dados1

        # Configura outros estilos, como bordas, fontes, etc.
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
        # Centralizar as colunas especificadas
        colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
        for coluna in colunas_centralizadas:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
         
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

    # def escrever_dados(self, sheet, start_row, dados):
    #     # Defina as colunas desejadas diretamente aqui
    #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

    #     # Escrever os dados
    #     for row, registro in enumerate(dados, start=start_row):
    #         for col, coluna in enumerate(colunas_desejadas, start=1):
    #             valor = getattr(registro, coluna, None)
    #             if coluna == 'valor_total' and valor is not None:
    #                 try:
    #                     valor = float(valor)
    #                 except ValueError:
    #                     valor = None
    #             sheet.cell(row=row, column=col).value = valor

    def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
        sheet = workbook.active
        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
        # total_linhas = linhas_dados1 + linhas_dados2

        
        # total_linhas = len(dados1) + len(dados2)
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        meses2 = {
            'JAN': 'janeiro',
            'FEV': 'fevereiro',
            'MAR': 'março',
            'ABR': 'abril',
            'MAI': 'maio',
            'JUN': 'junho',
            'JUL': 'julho',
            'AGO': 'agosto',
            'SET': 'setembro',
            'OUT': 'outubro',
            'NOV': 'novembro',
            'DEZ': 'dezembro'
        }

        self.mes2_texto = meses2.get(self.mes_texto)

        # Defina os estilos de borda
        dotted_border = Border(left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define as fontes
        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        # Define o preenchimento cinza claro para o cabeçalho
        gray_fill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')

        # Aplica os estilos para todas as células
        for row in sheet.iter_rows():
            for cell in row:
                # Aplica a borda
                cell.border = thin_border

                # Aplica a fonte padrão
                cell.font = fonte_padrao

        # Aplica estilos para o cabeçalho
        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # Aplica estilos para o cabeçalho 2
        for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # # Defina o estilo de borda externa
        # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                     top=Side(style='thin'), bottom=Side(style='thin'))
        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
        # for col in range(1, 9):
        #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
        #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
        # Seu texto
        texto = """MINISTERIO DA DEFESA
EXÉRCITO BRASILEIRO
COMANDO MILITAR DO OESTE
4ª BRIGADA DE CAVALARIA MECANIZADA
11º REGIMENTO DE CAVALARIA MECANIZADO
REGIMENTO MARECHAL DUTRA
"""

        # Insere o texto na célula A1
        sheet['A1'] = """MINISTERIO DA DEFESA"""
        sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
        sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
        sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
        sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
        sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
        sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
        sheet['A9'] = """HTM 1"""
        sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
        if total_htm1 is None:
            total_htm1 = decimal.Decimal('0.0')
        if total_htm2 is None:
            total_htm2 = decimal.Decimal('0.0')
        total_geral = total_htm1 + total_htm2
        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM PIX"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral

        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. PIX"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        # Define o idioma como português brasileiro
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        # Obter a data de hoje
        data_hoje = datetime.now()
        # Formatar a data como "dia de mês de ano"
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        # Inserir a data formatada na célula desejada
        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

        # Ajusta o alinhamento para envolver o texto e centralizá-lo
        sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

        # Mescla as células A1:H1
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

        # Define as células A7 e A9 com fonte negrito
        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



        # # tirando borda interna:

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        


# class RelatorioDinheiro(View):
#     nome_colunas_excel = {
#         'saida': 'Data de Saída',
#         'graduacao': 'PST/GRAD',
#         'nome_pagante': 'Hóspede',
#         'cpf_pagante': 'CPF',
#         'qtde_acomp': 'Nr Acomp',
#         'uh': 'UH',
#         'diarias': 'Dias',
#         'valor_total': 'Total (R$)',
#     }

#     cabecalho_fixo = 9
#     cabecalho_dados1 = 1
#     cabecalho_dados2 = 1
#     linhas_dados1 = None
#     linhas_dados2 = None
#     mes_texto = None
#     total_linhas = None
#     total_htm1 = 0
#     total_htm2 = 0
#     dados1 = None
#     dados2 = None
#     mes_relatorio = None
#     ano_relatorio = None


#     def get(self, request, *args, **kwargs):
#         self.mes_relatorio = request.GET.get('mes_relatorio')
#         self.ano_relatorio = request.GET.get('ano_relatorio')
#         forma_pagamento = request.GET.get('forma_pagamento')
#         # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
#         # print("Ano Relatório: auditoria", ano_relatorio)
#         # print("Forma Pagamento: auditoria", forma_pagamento)

#         meses = {
#             '1': 'JAN',
#             '2': 'FEV',
#             '3': 'MAR',
#             '4': 'ABR',
#             '5': 'MAI',
#             '6': 'JUN',
#             '7': 'JUL',
#             '8': 'AGO',
#             '9': 'SET',
#             '10': 'OUT',
#             '11': 'NOV',
#             '12': 'DEZ'
#         }

        
#         # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
#         if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
#             return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

#         # Validação dos dados de entrada
#         try:
#             self.mes_relatorio = int(self.mes_relatorio)
#             self.ano_relatorio = int(self.ano_relatorio)
#         except ValueError:
#             return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

#         if self.mes_relatorio < 1 or self.mes_relatorio > 12:
#             return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

#         if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
#             return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
#         self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
#         print(self.mes_texto)

#         # Realize as consultas usando os parâmetros recebidos
#         consultar_htm1 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_01",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).order_by('saida')

#         # Consulta para o segundo tipo de dados
#         consultar_htm2 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_02",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).order_by('saida')

#         if not consultar_htm1.exists() and not consultar_htm2.exists():
#             return render(request, 'portal/relatorio_pagamento_dinheiro.html', {'exibir_modal': True})

#         self.total_htm1 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_01",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).aggregate(soma=Sum('valor_total'))['soma']  

#         self.total_htm2 = BaseDados.objects.filter(
#             status_reserva="Pago",
#             mhex="HTM_02",
#             forma_pagamento=forma_pagamento,
#             saida__month=self.mes_relatorio,
#             saida__year=self.ano_relatorio
#         ).aggregate(soma=Sum('valor_total'))['soma']

#         # Combine os resultados das consultas em um único conjunto de dados
#         self.dados1 = list(consultar_htm1)
#         self.dados2 = list(consultar_htm2)

#         self.linhas_dados1 = len(self.dados1) 
#         self.linhas_dados2 = len(self.dados2)       
#         self.total_linhas = self.linhas_dados1 + self.linhas_dados2

#         # Gere a planilha
#         workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

#         # Configura o Excel
#         self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

#         # Configura o Estilo
#         self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

#         # Retorna a resposta HTTP com a planilha anexada
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

#         # Salva o workbook como uma resposta HTTP
#         workbook.save(response)
#         return response

#     def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
#         workbook = Workbook()
#         sheet = workbook.active
#         sheet.title = f'{ self.mes_texto}'

#         # INCLUIR CABEÇALHO DADOS 1
#         for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
#             sheet.cell(row=10, column=col).value = coluna

#         # INCLUIR DADOS 1
#         self.escrever_dados(sheet, 11, self.dados1)

#         # INCLUIR CABEÇALHO DADOS 2
#         for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
#             sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

#         # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
#         self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

#         return workbook

#     def escrever_dados(self, sheet, start_row, dados):
#         # Escrever os dados
#         for row, registro in enumerate(dados, start=start_row):
#             for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
#                 valor = getattr(registro, coluna, None)
#                 if coluna == 'valor_total' and valor is not None:
#                     try:
#                         valor = float(valor)
#                     except ValueError:
#                         valor = None
#                 elif coluna == 'graduacao':
#                 # Verifica se o status é diferente de "MILITAR DA ATIVA" ou "MILITAR DA RESERVA"
#                     if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
#                         valor = "CIVIL"        
#                 sheet.cell(row=row, column=col).value = valor

#     def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
#         # Acessa a planilha ativa
#         sheet = workbook.active

#         # Larguras das colunas desejadas
#         larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
#         for col, width in larguras_colunas.items():
#             sheet.column_dimensions[col].width = width

#         # Formatar coluna 'A' (Saida) para dd/mm/aaaa
#         date_format = NamedStyle(name='date_format')
#         date_format.number_format = 'DD/MM/YYYY'

#         # Aplicar estilo à coluna 'A' (Saida)
#         for cell in sheet['A']:
#             cell.style = date_format

#         # Ocultar as linhas de grade
#         sheet.sheet_view.showGridLines = False

#         # Formatar coluna 'H' (valor_total) para ##.###,##
#         valor_format = NamedStyle(name='valor_format')
#         valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

#         # Aplicar estilo à coluna 'H' (valor_total)
#         for cell in sheet['H']:
#             cell.style = valor_format

#         # Consulta para os dados do banco
#         consultar_dados = self.dados1

#         # Configura outros estilos, como bordas, fontes, etc.
#         self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
#         # Centralizar as colunas especificadas
#         colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
#         for coluna in colunas_centralizadas:
#             for cell in sheet[coluna]:
#                 cell.alignment = Alignment(horizontal='center')

#         # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
#         # dados1 = list(consultar_htm1)
#         # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
#         # dados2 = list(consultar_htm2)
#         # linhas_dados1 = len(dados1) 
#         # linhas_dados2 = len(dados2)       
         
#         sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
#         sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
#         sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

#     # def escrever_dados(self, sheet, start_row, dados):
#     #     # Defina as colunas desejadas diretamente aqui
#     #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

#     #     # Escrever os dados
#     #     for row, registro in enumerate(dados, start=start_row):
#     #         for col, coluna in enumerate(colunas_desejadas, start=1):
#     #             valor = getattr(registro, coluna, None)
#     #             if coluna == 'valor_total' and valor is not None:
#     #                 try:
#     #                     valor = float(valor)
#     #                 except ValueError:
#     #                     valor = None
#     #             sheet.cell(row=row, column=col).value = valor

#     def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
#         sheet = workbook.active
#         # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
#         # dados1 = list(consultar_htm1)
#         # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
#         # dados2 = list(consultar_htm2)
        
#         # linhas_dados1 = len(dados1) 
#         # linhas_dados2 = len(dados2)       
#         # total_linhas = linhas_dados1 + linhas_dados2

        
#         # total_linhas = len(dados1) + len(dados2)
#         sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
#         meses2 = {
#             'JAN': 'janeiro',
#             'FEV': 'fevereiro',
#             'MAR': 'março',
#             'ABR': 'abril',
#             'MAI': 'maio',
#             'JUN': 'junho',
#             'JUL': 'julho',
#             'AGO': 'agosto',
#             'SET': 'setembro',
#             'OUT': 'outubro',
#             'NOV': 'novembro',
#             'DEZ': 'dezembro'
#         }

#         self.mes2_texto = meses2.get(self.mes_texto)

#         # Defina os estilos de borda
#         dotted_border = Border(left=Side(style='dotted'),
#                             right=Side(style='dotted'),
#                             top=Side(style='dotted'),
#                             bottom=Side(style='dotted'))

#         thin_border = Border(left=Side(style='thin'),
#                             right=Side(style='thin'),
#                             top=Side(style='thin'),
#                             bottom=Side(style='thin'))

#         # Define as fontes
#         fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
#         fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

#         # Define o preenchimento cinza claro para o cabeçalho
#         gray_fill = PatternFill(start_color='00CCCCCC',
#                                 end_color='00CCCCCC',
#                                 fill_type='solid')

#         # Aplica os estilos para todas as células
#         for row in sheet.iter_rows():
#             for cell in row:
#                 # Aplica a borda
#                 cell.border = thin_border

#                 # Aplica a fonte padrão
#                 cell.font = fonte_padrao

#         # Aplica estilos para o cabeçalho
#         for row in sheet.iter_rows(min_row=10, max_row=10):
#             for cell in row:
#                 cell.font = fonte_negrito
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 cell.fill = gray_fill

#         # Aplica estilos para o cabeçalho 2
#         for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
#             for cell in row:
#                 cell.font = fonte_negrito
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                 cell.fill = gray_fill

#         # # Defina o estilo de borda externa
#         # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#         #                     top=Side(style='thin'), bottom=Side(style='thin'))
#         # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
#         # for col in range(1, 9):
#         #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
#         #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
#         #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
#         # Seu texto
#         texto = """MINISTERIO DA DEFESA
# EXÉRCITO BRASILEIRO
# COMANDO MILITAR DO OESTE
# 4ª BRIGADA DE CAVALARIA MECANIZADA
# 11º REGIMENTO DE CAVALARIA MECANIZADO
# REGIMENTO MARECHAL DUTRA
# """

#         # Insere o texto na célula A1
#         sheet['A1'] = """MINISTERIO DA DEFESA"""
#         sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
#         sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
#         sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
#         sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
#         sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
#         sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
#         sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
#         sheet['A9'] = """HTM 1"""
#         sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
#         sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
#         sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
#         total_htm1_decimal = decimal.Decimal(total_htm1 or 0)
#         total_htm2_decimal = decimal.Decimal(total_htm2 or 0)

#         total_geral = total_htm1_decimal + total_htm2_decimal
#         print("Valor de total_htm1:", total_htm1)
#         print("Valor de total_htm2:", total_htm2)
#         sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM DINHEIRO"
#         sheet.cell(row=total_linhas + 15, column=8).value = total_geral

#         sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
#         sheet.cell(row=total_linhas + 18, column=1).value = "DEP. DINHEIRO"
#         sheet.cell(row=total_linhas + 18, column=8).value = total_geral

#         # Define o idioma como português brasileiro
#         locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
#         # Obter a data de hoje
#         data_hoje = datetime.now()
#         # Formatar a data como "dia de mês de ano"
#         data_formatada = data_hoje.strftime("%d de %B de %Y")

#         # Inserir a data formatada na célula desejada
#         sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
#         sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
#         sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
#         sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
#         sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

#         # Ajusta o alinhamento para envolver o texto e centralizá-lo
#         sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
#         sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
#         sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

#         # Mescla as células A1:H1
#         sheet.merge_cells('A1:H1')
#         sheet.merge_cells('A2:H2')
#         sheet.merge_cells('A3:H3')
#         sheet.merge_cells('A4:H4')
#         sheet.merge_cells('A5:H5')
#         sheet.merge_cells('A6:H6')
#         sheet.merge_cells('A7:H7')
#         sheet.merge_cells('A8:H8')
#         sheet.merge_cells('A9:H9')
#         sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
#         sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
#         sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
#         sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

#         sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
#         sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

#         # Define as células A7 e A9 com fonte negrito
#         sheet['A7'].font = fonte_negrito
#         sheet['A9'].font = fonte_negrito
#         sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
#         sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
#         sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
#         sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


#         sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
#         sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
#         sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
#         sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
#         sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
#         sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
#         sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
#         sheet.cell(row=total_linhas + 14, column=8).border = thin_border
#         sheet.cell(row=total_linhas + 15, column=8).border = thin_border
#         sheet.cell(row=total_linhas + 17, column=1).border = thin_border
#         sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 8):
#         #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 9):
#         #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 8):
#         #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
#         # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
#         # for col in range(1, 8):
#         #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

#         # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
#         # for col in range(1, 9):
#         #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



#         # # tirando borda interna:

#         for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
#             for cell in row:
#                 if cell.row == 1:
#                     cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
#                 elif cell.row == 6:
#                     cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
#                 else:
#                     cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

#         for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
#             for cell in row:
#                 cell.border = None

#         for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
#             for cell in row:
#                 cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        






class RelatorioDinheiro(View):
    nome_colunas_excel = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    cabecalho_fixo = 9
    cabecalho_dados1 = 1
    cabecalho_dados2 = 1
    linhas_dados1 = None
    linhas_dados2 = None
    mes_texto = None
    total_linhas = None
    total_htm1 = 0
    total_htm2 = 0
    dados1 = None
    dados2 = None
    mes_relatorio = None
    ano_relatorio = None


    def get(self, request, *args, **kwargs):
        self.mes_relatorio = request.GET.get('mes_relatorio')
        self.ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
        # print("Ano Relatório: auditoria", ano_relatorio)
        # print("Forma Pagamento: auditoria", forma_pagamento)

        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        
        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

        # Validação dos dados de entrada
        try:
            self.mes_relatorio = int(self.mes_relatorio)
            self.ano_relatorio = int(self.ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if self.mes_relatorio < 1 or self.mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
        print(self.mes_texto)

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        # Consulta para o segundo tipo de dados
        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).order_by('saida')

        if not consultar_htm1.exists() and not consultar_htm2.exists():
            return HttpResponse("Não há lançamentos para o mês solicitado.")
        


        self.total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']  

        self.total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']

        # if not self.total_htm1.exists() and not self.total_htm2.exists():
        #     return HttpResponse("Não há lançamentos para o mês solicitado.")

        # Combine os resultados das consultas em um único conjunto de dados
        self.dados1 = list(consultar_htm1)
        self.dados2 = list(consultar_htm2)

        self.linhas_dados1 = len(self.dados1) 
        self.linhas_dados2 = len(self.dados2)       
        self.total_linhas = self.linhas_dados1 + self.linhas_dados2

        # Gere a planilha
        workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

        # Configura o Excel
        self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

        # Configura o Estilo
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

        # Retorna a resposta HTTP com a planilha anexada
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

        # Salva o workbook como uma resposta HTTP
        workbook.save(response)
        return response

    def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'{ self.mes_texto}'

        # INCLUIR CABEÇALHO DADOS 1
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=10, column=col).value = coluna

        # INCLUIR DADOS 1
        self.escrever_dados(sheet, 11, self.dados1)

        # INCLUIR CABEÇALHO DADOS 2
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

        # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
        self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

        return workbook

    def escrever_dados(self, sheet, start_row, dados):
        # Escrever os dados
        for row, registro in enumerate(dados, start=start_row):
            for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
                valor = getattr(registro, coluna, None)
                if coluna == 'valor_total' and valor is not None:
                    try:
                        valor = float(valor)
                    except ValueError:
                        valor = None
                elif coluna == 'graduacao':
                # Verifica se o status é diferente de "MILITAR DA ATIVA" ou "MILITAR DA RESERVA"
                    if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
                        valor = "CIVIL"         
                sheet.cell(row=row, column=col).value = valor

    def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
        # Acessa a planilha ativa
        sheet = workbook.active

        # Larguras das colunas desejadas
        larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras_colunas.items():
            sheet.column_dimensions[col].width = width

        # Formatar coluna 'A' (Saida) para dd/mm/aaaa
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'

        # Aplicar estilo à coluna 'A' (Saida)
        for cell in sheet['A']:
            cell.style = date_format

        # Ocultar as linhas de grade
        sheet.sheet_view.showGridLines = False

        # Formatar coluna 'H' (valor_total) para ##.###,##
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

        # Aplicar estilo à coluna 'H' (valor_total)
        for cell in sheet['H']:
            cell.style = valor_format

        # Consulta para os dados do banco
        consultar_dados = self.dados1

        # Configura outros estilos, como bordas, fontes, etc.
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
        # Centralizar as colunas especificadas
        colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
        for coluna in colunas_centralizadas:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
         
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

    # def escrever_dados(self, sheet, start_row, dados):
    #     # Defina as colunas desejadas diretamente aqui
    #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

    #     # Escrever os dados
    #     for row, registro in enumerate(dados, start=start_row):
    #         for col, coluna in enumerate(colunas_desejadas, start=1):
    #             valor = getattr(registro, coluna, None)
    #             if coluna == 'valor_total' and valor is not None:
    #                 try:
    #                     valor = float(valor)
    #                 except ValueError:
    #                     valor = None
    #             sheet.cell(row=row, column=col).value = valor

    def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
        sheet = workbook.active
        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
        # total_linhas = linhas_dados1 + linhas_dados2

        
        # total_linhas = len(dados1) + len(dados2)
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        meses2 = {
            'JAN': 'janeiro',
            'FEV': 'fevereiro',
            'MAR': 'março',
            'ABR': 'abril',
            'MAI': 'maio',
            'JUN': 'junho',
            'JUL': 'julho',
            'AGO': 'agosto',
            'SET': 'setembro',
            'OUT': 'outubro',
            'NOV': 'novembro',
            'DEZ': 'dezembro'
        }

        self.mes2_texto = meses2.get(self.mes_texto)

        # Defina os estilos de borda
        dotted_border = Border(left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define as fontes
        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        # Define o preenchimento cinza claro para o cabeçalho
        gray_fill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')

        # Aplica os estilos para todas as células
        for row in sheet.iter_rows():
            for cell in row:
                # Aplica a borda
                cell.border = thin_border

                # Aplica a fonte padrão
                cell.font = fonte_padrao

        # Aplica estilos para o cabeçalho
        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # Aplica estilos para o cabeçalho 2
        for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # # Defina o estilo de borda externa
        # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                     top=Side(style='thin'), bottom=Side(style='thin'))
        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
        # for col in range(1, 9):
        #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
        #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
        # Seu texto
        texto = """MINISTERIO DA DEFESA
EXÉRCITO BRASILEIRO
COMANDO MILITAR DO OESTE
4ª BRIGADA DE CAVALARIA MECANIZADA
11º REGIMENTO DE CAVALARIA MECANIZADO
REGIMENTO MARECHAL DUTRA
"""

        # Insere o texto na célula A1
        sheet['A1'] = """MINISTERIO DA DEFESA"""
        sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
        sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
        sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
        sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
        sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
        sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
        sheet['A9'] = """HTM 1"""
        sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
        total_htm1_decimal = decimal.Decimal(total_htm1 or 0)
        total_htm2_decimal = decimal.Decimal(total_htm2 or 0)
        total_geral = total_htm1_decimal + total_htm2_decimal
        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM DINHEIRO"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral

        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. DINHEIRO"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        # Define o idioma como português brasileiro
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        # Obter a data de hoje
        data_hoje = datetime.now()
        # Formatar a data como "dia de mês de ano"
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        # Inserir a data formatada na célula desejada
        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

        # Ajusta o alinhamento para envolver o texto e centralizá-lo
        sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

        # Mescla as células A1:H1
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

        # Define as células A7 e A9 com fonte negrito
        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



        # # tirando borda interna:

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        





# def obter_ocupacao_hotel(mes, ano, hotel):
#     ocupacao_hotel = {}
#     dias_no_mes = monthrange(ano, mes)[1]
    
#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}
#         for quarto in range(1, 13 if hotel == 'HTM_01' else 4):  # Determinar o número de quartos com base no hotel
#             # Verificar se há reservas para este dia e quarto
#             reservas = BaseDados.objects.filter(
#                 entrada__year=ano,
#                 entrada__month=mes,
#                 saida__year=ano,
#                 saida__month=mes,
#                 mhex=hotel,
#                 uh=str(quarto),
#                 entrada__day__lte=dia,
#                 saida__day__gte=dia
#             )
#             ocupacao_hotel[dia][quarto] = reservas.exists()
    
#     return ocupacao_hotel

# def visualizar_ocupacao_hotel(request):
#     mes = 3  # Janeiro
#     ano = 2024  # Ano desejado
#     ocupacao_htm_01 = obter_ocupacao_hotel(mes, ano, 'HTM_01')
#     ocupacao_htm_02 = obter_ocupacao_hotel(mes, ano, 'HTM_02')
    
#     return render(request, 'portal/ocupacao.html', {
#         'ocupacao_htm_01': ocupacao_htm_01,
#         'ocupacao_htm_02': ocupacao_htm_02,
#     })



# def obter_ocupacao_hotel(mes, ano, hotel):
#     ocupacao_hotel = {}
#     dias_no_mes = monthrange(ano, mes)[1]

#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}

#         for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
#             reservas_por_dia = BaseDados.objects.filter(
#                 saida__year=ano,
#                 saida__month=mes,
#                 mhex=hotel,
#                 saida__day=dia,  # Considerar apenas reservas com data de saída igual ao dia atual
#                 uh=str(quarto)  # Filtrar pelo número do quarto
#             )

#             if reservas_por_dia.exists():
#                 for reserva in reservas_por_dia:
#                     diarias = reserva.diarias
#                     valor_total = reserva.valor_total
#                     media_diaria = valor_total / diarias if diarias > 0 else 0
#                     for i in range(dia - diarias + 1, dia + 1):
#                         if i > 0:
#                             ocupacao_hotel[i][quarto] = {
#                                 'ocupado': True,
#                                 'media_diaria': "{:.2f}".format(media_diaria)  # Formatação decimal com duas casas decimais
#                             }
#             else:
#                 ocupacao_hotel[dia][quarto] = {
#                     'ocupado': False,
#                     'media_diaria': ""  # Deixando a string vazia para indicar que não há ocupação
#                 }

#     return ocupacao_hotel

######################################################################################################################################

# from decimal import Decimal
# from django.db.models import F, Sum, ExpressionWrapper, DecimalField
# from calendar import monthrange
# from datetime import datetime, timedelta


# def obter_ocupacao_hotel(mes, ano, hotel):
#     ocupacao_hotel = {}
#     dias_no_mes = monthrange(ano, mes)[1]
#     primeiro_dia_mes = datetime(ano, mes, 1).date()  # Primeiro dia do mês selecionado
#     ultimo_dia_mes = datetime(ano, mes, dias_no_mes).date()  # Último dia do mês selecionado

#     # Inicializar o dicionário para todos os dias e quartos
#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}
#         for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
#             ocupacao_hotel[dia][quarto] = {
#                 'ocupado': False,
#                 'diaria': ""  # Inicialmente vazio, indicando que não há reserva
#             }

#     # Consultar todas as reservas do hotel, independente do mês
#     reservas = BaseDados.objects.filter(
#         mhex=hotel,
#         status_reserva="Pago"
#     )

#     # Processar as reservas para calcular os dias de ocupação com base na entrada, saída e número de diárias
#     for reserva in reservas:
#         entrada = reserva.entrada  # Data de entrada
#         saida = reserva.saida  # Data de saída
        
#         # Verificar se entrada ou saída são None
#         if entrada is None or saida is None:
#             continue  # Pule esta reserva se algum desses campos estiver ausente
        
#         saida = saida - timedelta(days=1)  # Excluir o dia de saída
#         diarias = (saida - entrada).days + 1  # Número total de diárias da reserva
#         quarto = int(reserva.uh)
#         valor_dia = reserva.valor_dia  # Valor da diária 

#         # Calcular o valor adicional de pets (TOTAL_PET dividido por QTDE_PET e pelo número de dias da reserva)
#         total_pet = reserva.total_pet  # O valor de total_pet é garantido como Decimal pelo modelo
#         qtde_pet = reserva.qtde_pet if reserva.qtde_pet is not None else 0  # Garantir que qtde_pet não seja None
#         valor_pet_diaria = Decimal(0)  # Inicializamos como zero

#         if qtde_pet > 0:  # Evitar divisão por zero
#             valor_pet_diaria = total_pet / (qtde_pet * diarias)  # Valor proporcional por diária para pets

#         # Verificar se a reserva impacta o mês selecionado
#         if entrada <= ultimo_dia_mes and saida >= primeiro_dia_mes:
#             # Ajustar o início da ocupação para o mês selecionado, se a entrada for antes
#             inicio_ocupacao = max(entrada, primeiro_dia_mes)
#             fim_ocupacao = min(saida, ultimo_dia_mes)

#             # Marcar os dias de ocupação entre o início e o fim da ocupação ajustados
#             for dia_ocupacao in range(inicio_ocupacao.day, fim_ocupacao.day + 1):
#                 ocupacao_hotel[dia_ocupacao][quarto]['ocupado'] = True
                
#                 # Se já existe um valor na diária, somar com o novo valor
#                 if isinstance(ocupacao_hotel[dia_ocupacao][quarto]['diaria'], Decimal):
#                     ocupacao_hotel[dia_ocupacao][quarto]['diaria'] += valor_dia + valor_pet_diaria
#                 else:
#                     ocupacao_hotel[dia_ocupacao][quarto]['diaria'] = valor_dia + valor_pet_diaria

#     return ocupacao_hotel

from decimal import Decimal, ROUND_HALF_UP
from calendar import monthrange
from datetime import datetime, timedelta

def obter_ocupacao_hotel(mes, ano, hotel):
    ocupacao_hotel = {}
    dias_no_mes = monthrange(ano, mes)[1]
    primeiro_dia_mes = datetime(ano, mes, 1).date()  # Primeiro dia do mês selecionado
    ultimo_dia_mes = datetime(ano, mes, dias_no_mes).date()  # Último dia do mês selecionado

    # Inicializar o dicionário para todos os dias e quartos
    for dia in range(1, dias_no_mes + 1):
        ocupacao_hotel[dia] = {}
        for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
            ocupacao_hotel[dia][quarto] = {
                'ocupado': False,
                'diaria': Decimal('0.00')  # Inicialmente como Decimal, zero
            }

    # Consultar todas as reservas do hotel, independente do mês
    reservas = BaseDados.objects.filter(
        mhex=hotel,
        status_reserva="Pago"
    )

    # Processar as reservas para calcular os dias de ocupação com base na entrada, saída e número de diárias
    for reserva in reservas:
        entrada = reserva.entrada  # Data de entrada
        saida = reserva.saida  # Data de saída
        
        # Verificar se entrada ou saída são None
        if entrada is None or saida is None:
            continue  # Pule esta reserva se algum desses campos estiver ausente
        
        saida = saida - timedelta(days=1)  # Excluir o dia de saída
        diarias = (saida - entrada).days + 1  # Número total de diárias da reserva
        quarto = int(reserva.uh)
        valor_total = Decimal(str(reserva.valor_total))  # Garantir que é Decimal
        valor_dia = (valor_total / diarias).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

        # Verificar se a reserva impacta o mês selecionado
        if entrada <= ultimo_dia_mes and saida >= primeiro_dia_mes:
            # Ajustar o início da ocupação para o mês selecionado, se a entrada for antes
            inicio_ocupacao = max(entrada, primeiro_dia_mes)
            fim_ocupacao = min(saida, ultimo_dia_mes)

            # Marcar os dias de ocupação entre o início e o fim da ocupação ajustados
            for dia_ocupacao in range(inicio_ocupacao.day, fim_ocupacao.day + 1):
                ocupacao_hotel[dia_ocupacao][quarto]['ocupado'] = True
                
                # Somar ou atribuir o valor da diária
                ocupacao_hotel[dia_ocupacao][quarto]['diaria'] += valor_dia

    # Arredondar os valores finais de diária
    for dia in ocupacao_hotel:
        for quarto in ocupacao_hotel[dia]:
            ocupacao_hotel[dia][quarto]['diaria'] = ocupacao_hotel[dia][quarto]['diaria'].quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)

    return ocupacao_hotel






class Visualizar_Ocupacao_Hotel(View):
    def get(self, request, *args, **kwargs):
        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')

        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes e ano são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        mes_texto = meses.get(str(mes_relatorio), 'DESC')
        

        ocupacao_htm_01 = obter_ocupacao_hotel(mes_relatorio, ano_relatorio, 'HTM_01')
        ocupacao_htm_02 = obter_ocupacao_hotel(mes_relatorio, ano_relatorio, 'HTM_02')

        return render(request, 'portal/ocupacao_comvalores.html', {
            'ocupacao_htm_01': ocupacao_htm_01,
            'ocupacao_htm_02': ocupacao_htm_02,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,
            'mes_texto': mes_texto
        })

@login_required
def relatorio_ocupacao_finaceiro(request):    
    return render(request, 'portal/relatorio_ocupacao_financeira.html')         


#############################################################################

def obter_censo_mes(mes, ano, hotel):
    relatorio = []
    dias_no_mes = monthrange(ano, mes)[1]
    primeiro_dia_mes = datetime(ano, mes, 1).date()
    ultimo_dia_mes = datetime(ano, mes, dias_no_mes).date()

    # Inicializar o dicionário para todos os dias com campos QTDE e VALOR
    for dia in range(1, dias_no_mes + 1):
        relatorio.append({
            'dia': dia,
            'qtde': 0,  # Inicializar a quantidade
            'valor': 0.0  # Inicializar o valor total do dia
        })

    # Processar reservas do hotel específico (HTM_01 ou HTM_02)
    reservas = BaseDados.objects.filter(
        mhex=hotel,  # Filtrar para o hotel específico
        status_reserva="Pago"
    )

    # Iterar sobre cada dia do mês
    for dia in range(1, dias_no_mes + 1):
        data_atual = datetime(ano, mes, dia).date()

        # Verificar as reservas ativas no dia atual (considerando entrada e saída)
        reservas_dia = reservas.filter(entrada__lte=data_atual, saida__gt=data_atual)

        # Para cada reserva do dia, acumular qtde e valor
        for reserva in reservas_dia:
            qtde_hosp = reserva.qtde_hosp or 0  # Usar o campo QTDE_HOSP para calcular a quantidade de hóspedes
            valor_dia = float(reserva.valor_dia)  # Usar diretamente o valor diário do banco de dados

            # Somar a quantidade de hóspedes e o valor para o dia
            relatorio[dia - 1]['qtde'] += qtde_hosp  # Adiciona o número de hóspedes por dia
            relatorio[dia - 1]['valor'] += valor_dia  # Somar o valor diário corretamente

            # Depuração: Imprimir os valores sendo somados para o dia 1
            if dia == 1:
                print(f"Dia {dia} - Hotel {hotel} - Reserva ID: {reserva.id} | Valor por dia: {valor_dia} | Qtde Hospedes: {qtde_hosp} | Valor Total Somado: {relatorio[dia - 1]['valor']}")

    return relatorio

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.views import View

from openpyxl import Workbook
from django.http import HttpResponse

class Visualizar_Censo_Mes(View):
    def get(self, request, *args, **kwargs):
        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        # Obter o ano selecionado pelo usuário
        ano_relatorio = request.GET.get('ano_relatorio')

        if not ano_relatorio:
            return render(request, 'portal/relatorio_censo_mes.html', {
                'error_message': "Por favor, selecione o ano para gerar o relatório."
            })

        # Validar o ano
        try:
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Ano deve ser um número inteiro.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Obter os dados de censo para todos os meses do ano
        censo_htm_01 = obter_censo_ano(ano_relatorio, 'HTM_01')
        censo_htm_02 = obter_censo_ano(ano_relatorio, 'HTM_02')

        # Criar um workbook e uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Censo"

        # Definir cabeçalhos
        ws.append(["Mês", "HTM_01 - Quantidade de Hóspedes", "HTM_01 - Valor", "HTM_02 - Quantidade de Hóspedes", "HTM_02 - Valor"])

        # Adicionar dados para cada mês
        for mes in range(1, 13):
            mes_nome = meses.get(str(mes), 'Desconhecido')
            censo_01 = censo_htm_01[mes - 1]  # Dados HTM_01
            censo_02 = censo_htm_02[mes - 1]  # Dados HTM_02

            # Para cada mês, pegar os dados de quantidade e valor (ajuste conforme seu modelo de dados)
            qtde_htm_01 = sum([item.qtde for item in censo_01])
            valor_htm_01 = sum([item.valor for item in censo_01])

            qtde_htm_02 = sum([item.qtde for item in censo_02])
            valor_htm_02 = sum([item.valor for item in censo_02])

            # Adicionar linha ao Excel
            ws.append([mes_nome, qtde_htm_01, valor_htm_01, qtde_htm_02, valor_htm_02])

        # Preparar a resposta para download do Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="relatorio_censo_{ano_relatorio}.xlsx"'
        wb.save(response)

        return response

from django.http import HttpResponse
from django.views import View
from datetime import datetime
import pandas as pd

def obter_censo_ano(ano_relatorio, hotel):
    censo_ano = []
    for mes in range(1, 13):  # Para todos os meses de janeiro a dezembro
        censo_mes = Censo.objects.filter(ano=ano_relatorio, mes=mes, hotel=hotel)
        censo_ano.append(censo_mes)
    return censo_ano

class Visualizar_Censo_Ano(View):
    def get(self, request):
        # Captura o ano informado ou usa o ano atual como padrão
        ano = request.GET.get('ano', datetime.now().year)
        
        # Filtra as reservas pelo ano selecionado
        reservas = Reserva.objects.filter(data__year=ano)
        
        # Organiza os dados para exportação
        dados = reservas.values('id', 'data', 'total_consumacao', 'total_pet', 'qtde_pet')
        df = pd.DataFrame.from_records(dados)

        # Adiciona resumo ao final do Excel
        resumo = {
            'Total Consumo': reservas.aggregate(total_consumo=models.Sum('total_consumacao'))['total_consumo'] or 0,
            'Total PET': reservas.aggregate(total_pet=models.Sum('total_pet'))['total_pet'] or 0,
            'Quantidade PET': reservas.aggregate(qtde_pet=models.Sum('qtde_pet'))['qtde_pet'] or 0,
        }
        df_resumo = pd.DataFrame([resumo])

        # Cria um arquivo Excel com duas abas: Dados e Resumo
        with pd.ExcelWriter('relatorio_censo_ano.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Dados')
            df_resumo.to_excel(writer, index=False, sheet_name='Resumo')

        # Configura a resposta HTTP com o arquivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=relatorio_censo_{ano}.xlsx'
        writer.book.save(response)
        return response






@login_required
def relatorio_censo_mes(request):    
    return render(request, 'portal/relatorio_censo_mes.html')   

#############################################################################

def obter_censo_ano_excel(ano, hotel):
    # Criar o workbook e a primeira planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Relatório {ano}"

    # Meses em formato de texto para o cabeçalho
    meses_nome = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 
                  'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']

    # Adicionar o cabeçalho de dois níveis
    ws.merge_cells('A1:A2')
    ws['A1'] = 'DIA'
    
    col = 2
    for i, mes in enumerate(meses_nome):
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        ws.cell(row=1, column=col, value=mes).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col, value='QTDE')
        ws.cell(row=2, column=col+1, value='VALOR')
        col += 2

    # Preencher os dados para cada dia
    for dia in range(1, 32):  # Dias de 1 a 31
        linha = [dia]  # Primeira coluna é o dia
        ws.cell(row=dia+2, column=1, value=dia)  # Preencher a coluna dos dias

        col = 2
        for mes in range(1, 13):  # Para cada mês, preencher QTDE e VALOR
            # Obter o censo para o mês e hotel
            censo_mes = obter_censo_mes(mes, ano, hotel)

            # Verificar se o dia existe no mês (pode haver menos de 31 dias em alguns meses)
            if dia <= len(censo_mes):
                # Adicionar QTDE e VALOR
                qtde = censo_mes[dia - 1]['qtde']
                valor = censo_mes[dia - 1]['valor']
                ws.cell(row=dia+2, column=col, value=qtde)
                ws.cell(row=dia+2, column=col+1, value=valor)
            else:
                # Se o dia não existir no mês, adicionar valores vazios
                ws.cell(row=dia+2, column=col, value='-')
                ws.cell(row=dia+2, column=col+1, value='-')
            col += 2

    # Preparar a resposta HTTP com o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{hotel}_{ano}.xlsx"'

    # Salvar o arquivo no response
    wb.save(response)
    return response

from django.http import HttpResponse
from django.views import View

class Visualizar_Censo_Ano_Excel(View):
    def get(self, request, *args, **kwargs):
        # Verificar se o parâmetro ano_relatorio está presente
        ano_relatorio = request.GET.get('ano_relatorio')

        if not ano_relatorio:
            return HttpResponse("Ano do relatório não foi fornecido.", status=400)

        # Validação do ano
        try:
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponse("Ano deve ser um número inteiro.", status=400)

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponse("Ano deve estar entre 2024 e 2030.", status=400)

        # Obter o hotel, aqui vamos assumir que está sendo passado via query params, ou fixar para testes
        hotel = request.GET.get('hotel', 'HTM_01')  # Default para HTM_01, você pode ajustar isso

        # Gerar o Excel para o hotel e ano
        return obter_censo_ano_excel(ano_relatorio, hotel)

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment










###############################################################################


from datetime import datetime, timedelta
from collections import OrderedDict
from calendar import monthrange
from django.db.models import Q

def obter_agenda(mes, ano, hotel):
    ocupacao_hotel = OrderedDict()
    dias_no_mes = monthrange(ano, mes)[1]

    # Inicializar todos os dias do mês para cada quarto
    for dia in range(1, dias_no_mes + 1):
        ocupacao_hotel[dia] = {}
        for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
            ocupacao_hotel[dia][str(quarto)] = {
                'ocupado': False,
                'media_diaria': ""
            }

    # Consultar todas as reservas, excluindo status "Pendente" ou "Recusada"
    reservas = BaseDados.objects.filter(
        mhex=hotel,
        entrada__year=ano,
        entrada__month=mes,
        status_reserva__in=['Confirmada', 'Pago']  # Certifique-se de que esses são os status corretos
    )

    for reserva in reservas:
        diarias = reserva.diarias if reserva.diarias > 0 else 0
        quarto = str(reserva.uh)
        valor_total = reserva.valor_total
        media_diaria = valor_total / diarias if diarias > 0 else 0

        # Calcular e marcar os dias da reserva como ocupados
        data_entrada = reserva.entrada
        data_saida = data_entrada + timedelta(days=diarias)

        while data_entrada < data_saida:
            dia_reserva = data_entrada.day
            if dia_reserva <= dias_no_mes:
                if quarto in ocupacao_hotel[dia_reserva]:
                    ocupacao_hotel[dia_reserva][quarto]['ocupado'] = True
                    ocupacao_hotel[dia_reserva][quarto]['media_diaria'] = "{:.2f}".format(media_diaria)
            data_entrada += timedelta(days=1)

    # Ordenar o dicionário por dia
    ocupacao_hotel = OrderedDict(sorted(ocupacao_hotel.items()))

    return ocupacao_hotel




# def obter_agenda(mes, ano, hotel):
#     ocupacao_hotel = OrderedDict()
#     dias_no_mes = monthrange(ano, mes)[1]

#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}

#         for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
#             reservas_por_dia = BaseDados.objects.filter(
#                 mhex=hotel,
#                 entrada__lte=datetime(ano, mes, dia).date(),  # Considerar reservas com data de entrada menor ou igual ao dia atual
#                 saida__gte=datetime(ano, mes, dia).date(),  # Considerar reservas com data de saída maior ou igual ao dia atual
#                 uh=str(quarto)  # Filtrar pelo número do quarto
#             )

#             if reservas_por_dia.exists():
#                 for reserva in reservas_por_dia:
#                     data_saida = reserva.entrada + timedelta(days=reserva.diarias - 1)  # Data de saída do hóspede
#                     if reserva.entrada <= datetime(ano, mes, dia).date() <= data_saida:
#                         diarias_ocupadas = min(reserva.diarias, (data_saida - reserva.entrada).days + 1)
#                         for i in range(diarias_ocupadas):
#                             dia_atual = (reserva.entrada + timedelta(days=i)).day
#                             if dia_atual not in ocupacao_hotel:
#                                 ocupacao_hotel[dia_atual] = {}
#                             if quarto not in ocupacao_hotel[dia_atual]:
#                                 ocupacao_hotel[dia_atual][quarto] = {}
#                             ocupacao_hotel[dia_atual][quarto] = {
#                                 'ocupado': True,
#                                 'media_diaria': "{:.2f}".format(reserva.valor_total / reserva.diarias)
#                             }
#                     else:
#                         if quarto not in ocupacao_hotel[dia]:
#                             ocupacao_hotel[dia][quarto] = {}
#                         ocupacao_hotel[dia][quarto] = {
#                             'ocupado': False,
#                             'media_diaria': ""
#                         }
#             else:
#                 if quarto not in ocupacao_hotel[dia]:
#                     ocupacao_hotel[dia][quarto] = {}
#                 ocupacao_hotel[dia][quarto] = {
#                     'ocupado': False,
#                     'media_diaria': ""
#                 }

#     # Ordenar o dicionário por dia
#     ocupacao_hotel = OrderedDict(sorted(ocupacao_hotel.items()))

#     return ocupacao_hotel

# def obter_agenda2(mes, ano, hotel):
#     ocupacao_hotel = OrderedDict()
#     dias_no_mes = monthrange(ano, mes)[1]

#     for dia in range(1, dias_no_mes + 1):
#         ocupacao_hotel[dia] = {}

#         for quarto in range(1, 13 if hotel == 'HTM_01' else 5):
#             # Inicializar o dicionário para o quarto no dia especificado
#             ocupacao_hotel[dia][quarto] = {
#                 'ocupado': False,
#                 'total_diaria': 0.0,  # Iniciar totalizador de diária
#                 'contagem_reservas': 0  # Contar reservas para esse quarto nesse dia
#             }

#             reservas_por_dia = BaseDados.objects.filter(
#                 mhex=hotel,
#                 entrada__lte=datetime(ano, mes, dia).date(),
#                 saida__gte=datetime(ano, mes, dia).date(),
#                 uh=str(quarto), 
#                 status_reserva='Pago' # Filtrar pelo número do quarto
#             )

#             for reserva in reservas_por_dia:
#                 ocupacao_hotel[dia][quarto]['ocupado'] = True
#                 ocupacao_hotel[dia][quarto]['total_diaria'] += reserva.valor_total / reserva.diarias
#                 ocupacao_hotel[dia][quarto]['contagem_reservas'] += 1

#             # Calcular a média diária após processar todas as reservas para aquele quarto naquele dia
#             if ocupacao_hotel[dia][quarto]['contagem_reservas'] > 0:
#                 ocupacao_hotel[dia][quarto]['media_diaria'] = "{:.2f}".format(
#                     ocupacao_hotel[dia][quarto]['total_diaria'] / ocupacao_hotel[dia][quarto]['contagem_reservas']
#                 )
#             else:
#                 # Se não houver reservas, defina a média diária como vazia
#                 ocupacao_hotel[dia][quarto]['media_diaria'] = ""

#     # Ordenar o dicionário por dia pode ser opcional se a ordem dos dias já é gerada corretamente
#     return ocupacao_hotel


class Visualizar_Agenda_Ocupacao(View):
    def get(self, request, *args, **kwargs):
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')

        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes e ano são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        ocupacao_htm_01 = obter_agenda(mes_relatorio, ano_relatorio, 'HTM_01')
        ocupacao_htm_02 = obter_agenda(mes_relatorio, ano_relatorio, 'HTM_02')

        return render(request, 'portal/ocupacao.html', {
            'ocupacao_htm_01': ocupacao_htm_01,
            'ocupacao_htm_02': ocupacao_htm_02,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,
        })

@login_required
def agenda_ocupacao(request):    
    return render(request, 'portal/relatorio_agenda.html')  


@login_required
def relatorio_todos(request):    
    return render(request, 'portal/relatorio_todos.html') 


@login_required
def consultar_todos(request):
    if request.method == 'GET':
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')     

        # Armazenar os parâmetros na sessão
        request.session['mes_relatorio'] = mes_relatorio
        request.session['ano_relatorio'] = ano_relatorio   
        
        # Verificar se mes_relatorio e ano_relatorio não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio e ano_relatorio são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")
        
        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")
        
        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(            
            saida__month=mes_relatorio,
            saida__year=ano_relatorio,
            mhex="HTM_01",
        ).exclude(status_reserva="Recusada").order_by('entrada')      

          

        consultar_htm2 = BaseDados.objects.filter(            
            saida__month=mes_relatorio,
            saida__year=ano_relatorio,
            mhex="HTM_02",
        ).exclude(status_reserva="Recusada").order_by('entrada')

        

        # Atualização em lote dos objetos BaseDados
        if consultar_htm1.exists():
            consultar_htm1.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        if consultar_htm2.exists():
            consultar_htm2.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        context = {
            'consultar_htm1': consultar_htm1,
            'consultar_htm2': consultar_htm2,
            'mes_relatorio': mes_relatorio,
            'ano_relatorio': ano_relatorio,           

        }
        return render(request, 'portal/consultar_todos.html', context)
    else:
        return HttpResponseBadRequest("Método de requisição inválido.")


@login_required
def editar_todos(request, reservas_pk):
    # Aqui usamos get_object_or_404 para garantir que se o objeto não for encontrado,
    # uma página de erro 404 será retornada.
    editar = get_object_or_404(BaseDados, pk=reservas_pk)
    mes_relatorio = request.session.get('mes_relatorio')
    ano_relatorio = request.session.get('ano_relatorio')
    
    
    # Se estamos lidando com uma requisição POST, significa que o formulário foi submetido
    if request.method == 'POST':
        # Ajuste do valor_ajuste
        if 'valor_ajuste' in request.POST:
            valor_ajuste = request.POST['valor_ajuste'].replace(',', '.')
            request.POST = request.POST.copy()
            request.POST['valor_ajuste'] = valor_ajuste
        
        # Passamos a instância da reserva que queremos editar para o formulário,
        # junto com os dados submetidos (request.POST)
        form = ReservasForm(request.POST, instance=editar)
        print("Dados recebidos do formulário via POST:")
        print(request.POST)
        if 'status_reserva' in request.POST:
            status_reserva = request.POST['status_reserva']
            print("Status da reserva recebido:", status_reserva)
            # Atribuir o valor do campo status_reserva à instância do formulário
            form.instance.status_reserva = status_reserva
        
        # Verificamos se o formulário é válido
        if form.is_valid():
            print("Formulário válido. Tentando salvar...")
            # Se for válido, salvamos as alterações feitas na reserva
            form.save()
            print("Reserva salva com sucesso!")
            # Construir a URL de redirecionamento para consultar_todos com os parâmetros de consulta
            redirect_url = reverse('consultar_todos')
            redirect_url += f'?mes_relatorio={mes_relatorio}&ano_relatorio={ano_relatorio}'
            # Redirecionar para consultar_todos com os parâmetros de consulta
            return redirect(redirect_url)
        else:
            print("Formulário inválido. Erros de validação:")
            print(form.errors)
    else:
        # Se a requisição for um GET, criamos um formulário preenchido com as informações
        # da reserva que queremos editar
        form = ReservasForm(instance=editar)
    
    # Passamos o formulário preenchido para o template
    context = {
        'form': form, 
        'reserva': editar,
        'mes_relatorio': mes_relatorio,
        'ano_relatorio': ano_relatorio, 
    }
    return render(request, 'portal/editar_todos.html', context)



from collections import defaultdict
from datetime import datetime
from calendar import monthrange

from collections import defaultdict
from calendar import monthrange
from datetime import date
from .models import BaseDados  # Ajuste conforme o caminho correto do seu modelo

def obter_relatorio_hospedes_valor(mes, ano, hotel):
    relatorio = defaultdict(lambda: {'total_hospedes': 0, 'valor_total': Decimal('0.0')})
    dias_no_mes = monthrange(ano, mes)[1]

    print(f"Verificando hotel: {hotel}, mês: {mes}, ano: {ano}")

    for dia in range(1, dias_no_mes + 1):
        data_atual = date(ano, mes, dia)

        reservas_ativas_no_dia = BaseDados.objects.filter(
            entrada__lte=data_atual,
            saida__gte=data_atual,
            mhex=hotel
        )

        print(f"Dia {dia}: encontradas {reservas_ativas_no_dia.count()} reservas ativas")

        for reserva in reservas_ativas_no_dia:
            qtde_hosp = Decimal(reserva.qtde_hosp if reserva.qtde_hosp > 0 else 1)
            valor_por_pessoa = Decimal(reserva.valor_total) / qtde_hosp

            relatorio[dia]['total_hospedes'] += qtde_hosp
            relatorio[dia]['valor_total'] += valor_por_pessoa * qtde_hosp

            print(f"Reserva ID {reserva.id}: qtde_hosp={qtde_hosp}, valor_total={reserva.valor_total}")

    # Preparação para visualização dos resultados acumulados
    for dia, dados in relatorio.items():
        print(f"Dia {dia}: total_hospedes={dados['total_hospedes']}, valor_total={dados['valor_total']}")

    relatorio_formatado = {
        dia: {
            'total_hospedes': relatorio[dia]['total_hospedes'],
            'valor_total': "{:.2f}".format(relatorio[dia]['valor_total'])
        } for dia in relatorio
    }

    return relatorio_formatado  


def relatorio_anual_hoteis(request):
    ano = 2024
    hoteis = ['HTM_01', 'HTM_02']
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    relatorios = []

    # Criar uma lista de dias para cada mês (assumindo que todos os meses têm 31 dias)
    dias_no_mes = list(range(1, 32))

    for hotel in hoteis:
        dados_hotel = {'nome': hotel, 'dados_meses': [], 'dias_no_mes': dias_no_mes}
        for mes_idx, mes in enumerate(meses, start=1):
            # Usando a função correta para recuperar os dados
            dados_mes = obter_relatorio_hospedes_valor(mes_idx, ano, hotel)
            dados_hotel['dados_meses'].append({
                'mes': mes,
                'total_hospedes': dados_mes.get('total_hospedes', {}),
                'valor_total': dados_mes.get('valor_total', {})
            })
        relatorios.append(dados_hotel)

    context = {
        'relatorios': relatorios,
        'ano': ano,
    }
    return render(request, 'portal/consultar_censo1.html', context)
