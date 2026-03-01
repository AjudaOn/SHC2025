from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from sqlalchemy import extract, func
from models import db, BaseDados
from datetime import datetime
from decimal import Decimal
import decimal
import locale

class ReportService:
    """
    Service to handle Excel report generation.
    """
    
    NOME_COLUNAS_EXCEL = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    MESES = {
        '1': 'JAN', '2': 'FEV', '3': 'MAR', '4': 'ABR',
        '5': 'MAI', '6': 'JUN', '7': 'JUL', '8': 'AGO',
        '9': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'
    }

    @staticmethod
    def generate_payment_report(mes_relatorio, ano_relatorio, forma_pagamento):
        """
        Generates an Excel workbook for the payment report.
        Returns the workbook object or raises ValueError on validation error.
        """
        
        # Validation
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except (ValueError, TypeError):
            raise ValueError("Mês e ano devem ser números inteiros.")

        if not (1 <= mes_relatorio <= 12):
            raise ValueError("Mês deve estar entre 1 e 12.")
        
        if not (2024 <= ano_relatorio <= 2030):
            raise ValueError("Ano deve estar entre 2024 e 2030.")

        if not forma_pagamento:
            raise ValueError("Forma de pagamento é obrigatória.")

        mes_texto = ReportService.MESES.get(str(mes_relatorio), 'DESC')

        # Data Fetching using SQLAlchemy
        def fetch_data(mhex):
            return BaseDados.query.filter(
                BaseDados.status_reserva == "Pago",
                BaseDados.mhex == mhex,
                BaseDados.forma_pagamento == forma_pagamento,
                extract('month', BaseDados.saida) == mes_relatorio,
                extract('year', BaseDados.saida) == ano_relatorio
            ).order_by(BaseDados.saida).all()

        dados1 = fetch_data("HTM_01")
        dados2 = fetch_data("HTM_02")

        if not dados1 and not dados2:
            return None

        # Totals using SQLAlchemy
        def get_total(mhex):
            result = db.session.query(func.sum(BaseDados.valor_total)).filter(
                BaseDados.status_reserva == "Pago",
                BaseDados.mhex == mhex,
                BaseDados.forma_pagamento == forma_pagamento,
                extract('month', BaseDados.saida) == mes_relatorio,
                extract('year', BaseDados.saida) == ano_relatorio
            ).scalar()
            return result or 0

        total_htm1 = get_total("HTM_01")
        total_htm2 = get_total("HTM_02")

        # Workbook Generation
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = mes_texto

        linhas_dados1 = len(dados1)
        linhas_dados2 = len(dados2)

        # Headers and Data
        ReportService._write_section(sheet, 10, dados1)
        ReportService._write_section(sheet, linhas_dados1 + 13, dados2)

        # Configuration and Styles
        ReportService._configure_excel(sheet, ano_relatorio)
        total_linhas = linhas_dados1 + linhas_dados2
        ReportService._configure_styles(
            sheet,
            dados1,
            dados2,
            linhas_dados1,
            linhas_dados2,
            total_linhas,
            total_htm1,
            total_htm2,
            ano_relatorio,
            mes_texto
        )

        return workbook

    @staticmethod
    def _write_section(sheet, start_row, dados):
        # Headers
        for col, coluna in enumerate(ReportService.NOME_COLUNAS_EXCEL.values(), start=1):
            sheet.cell(row=start_row, column=col).value = coluna
        
        # Data
        for row, registro in enumerate(dados, start=start_row + 1):
            for col, key in enumerate(ReportService.NOME_COLUNAS_EXCEL.keys(), start=1):
                valor = getattr(registro, key, None)

                if key == 'nome_pagante' and not valor:
                    valor = getattr(registro, 'nome', None)
                elif key == 'cpf_pagante' and not valor:
                    valor = getattr(registro, 'cpf', None)
                elif key == 'qtde_acomp' and valor is None:
                    qtde_hosp = getattr(registro, 'qtde_hosp', None)
                    valor = max((qtde_hosp or 0) - 1, 0)
                elif key == 'diarias' and valor is None:
                    valor = 0
                elif key == 'uh' and valor is None:
                    valor = ''
                elif key == 'valor_total':
                    if valor is None:
                        valor = 0.0
                    else:
                        try:
                            valor = float(valor)
                        except (ValueError, TypeError):
                            valor = 0.0
                elif key == 'graduacao':
                    if registro.status not in ["MILITAR DA ATIVA", "MILITAR DA RESERVA"]:
                        valor = "CIVIL"
                
                sheet.cell(row=row, column=col).value = valor

    @staticmethod
    def _configure_excel(sheet, ano_relatorio):
        # Column Widths
        larguras = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras.items():
            sheet.column_dimensions[col].width = width

        # Date Format
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'
        existing_styles = [s.name if hasattr(s, 'name') else s for s in sheet.parent.named_styles]
        if 'date_format' not in existing_styles:
            sheet.parent.add_named_style(date_format)
        for cell in sheet['A']:
            cell.style = 'date_format'

        # Value Format
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00'
        existing_styles = [s.name if hasattr(s, 'name') else s for s in sheet.parent.named_styles]
        if 'valor_format' not in existing_styles:
            sheet.parent.add_named_style(valor_format)
        for cell in sheet['H']:
            cell.style = 'valor_format'

        sheet.sheet_view.showGridLines = False

        # Centralizar colunas específicas
        for coluna in ['A', 'B', 'D', 'E', 'F', 'G']:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

    @staticmethod
    def _configure_styles(sheet, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio, mes_texto):
        meses2 = {
            'JAN': 'janeiro', 'FEV': 'fevereiro', 'MAR': 'março', 'ABR': 'abril',
            'MAI': 'maio', 'JUN': 'junho', 'JUL': 'julho', 'AGO': 'agosto',
            'SET': 'setembro', 'OUT': 'outubro', 'NOV': 'novembro', 'DEZ': 'dezembro'
        }
        mes2_texto = meses2.get(mes_texto, '')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        gray_fill = PatternFill(start_color='00CCCCCC', end_color='00CCCCCC', fill_type='solid')

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.font = fonte_padrao

        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        header2_row = len(dados1) + 11 + 2
        for row in sheet.iter_rows(min_row=header2_row, max_row=header2_row):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        sheet['A1'] = "MINISTERIO DA DEFESA"
        sheet['A2'] = "EXÉRCITO BRASILEIRO"
        sheet['A3'] = "COMANDO MILITAR DO OESTE"
        sheet['A4'] = "4ª BRIGADA DE CAVALARIA MECANIZADA"
        sheet['A5'] = "11º REGIMENTO DE CAVALARIA MECANIZADO"
        sheet['A6'] = "REGIMENTO MARECHAL DUTRA"
        sheet['A7'] = "RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {mes2_texto} de {ano_relatorio}."
        sheet['A9'] = "HTM 1"

        sheet.cell(row=linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=linhas_dados1 + 12, column=1).value = "HTM 2"
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2

        if total_htm1 is None:
            total_htm1 = decimal.Decimal('0.0')
        if total_htm2 is None:
            total_htm2 = decimal.Decimal('0.0')
        total_geral = total_htm1 + total_htm2

        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM PIX"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral
        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. PIX"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        try:
            locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        except locale.Error:
            try:
                locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')
            except locale.Error:
                pass
        data_hoje = datetime.now()
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        for row_id in range(1, 9):
            sheet[f'A{row_id}'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)

        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito

        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=total_linhas + 19, max_row=total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=total_linhas + 16, max_row=total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))

    @staticmethod
    def generate_censo_ano_excel(ano, hotel, get_censo_mes_fn=None, censo_por_mes=None):
        if censo_por_mes is None:
            if not get_censo_mes_fn:
                raise ValueError("get_censo_mes_fn é obrigatório quando censo_por_mes não é fornecido.")
            censo_por_mes = {mes: get_censo_mes_fn(mes, ano, hotel) for mes in range(1, 13)}

        wb = Workbook()
        ws = wb.active
        ws.title = f"Censo {ano}"

        fonte_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        fonte_padrao = Font(name="Calibri", size=10)
        fonte_negrito = Font(name="Calibri", size=10, bold=True)
        alinhamento_central = Alignment(horizontal="center", vertical="center")
        fundo_titulo = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        fundo_dia = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")
        fundo_mes_claro = PatternFill(start_color="F2F8EE", end_color="F2F8EE", fill_type="solid")
        fundo_mes_escuro = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")

        borda_titulo = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        borda_dia = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        borda_mes = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style=None)
        )
        borda_qtd = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )
        borda_valor = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )
        borda_ext = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        borda_int = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        borda_total = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )

        for start_col in range(2, 26, 2):
            for row in range(4, 35):
                for col in range(start_col, start_col + 2):
                    if row == 4:
                        ws.cell(row=row, column=col).border = Border(
                            left=borda_ext.left if col == start_col else borda_int.left,
                            right=borda_int.right if col == start_col else borda_ext.right,
                            top=borda_ext.top,
                            bottom=borda_int.bottom
                        )
                    elif row == 34:
                        ws.cell(row=row, column=col).border = Border(
                            left=borda_ext.left if col == start_col else borda_int.left,
                            right=borda_int.right if col == start_col else borda_ext.right,
                            top=borda_int.top,
                            bottom=borda_ext.bottom
                        )
                    else:
                        ws.cell(row=row, column=col).border = Border(
                            left=borda_ext.left if col == start_col else borda_int.left,
                            right=borda_int.right if col == start_col else borda_ext.right,
                            top=borda_int.top,
                            bottom=borda_int.bottom
                        )

        ws.merge_cells('A1:Y1')
        titulo_cell = ws['A1']
        titulo_cell.value = "HÓSPEDES E VALORES ARRECADADOS"
        titulo_cell.font = fonte_titulo
        titulo_cell.alignment = alinhamento_central
        titulo_cell.fill = fundo_titulo
        titulo_cell.border = borda_titulo

        for col in range(1, 26):
            ws.cell(row=1, column=col).border = borda_titulo

        ws.cell(row=35, column=1).border = borda_total

        for start_col in range(2, 26, 2):
            for col in range(start_col, start_col + 2):
                ws.cell(row=35, column=col).border = Border(
                    left=borda_ext.left if col == start_col else borda_int.left,
                    right=borda_int.right if col == start_col else borda_ext.right,
                    top=borda_int.top,
                    bottom=borda_ext.bottom
                )

        ws.merge_cells('A2:A3')
        ws['A2'] = "DIA"
        ws['A2'].font = fonte_negrito
        ws['A2'].alignment = alinhamento_central
        ws['A2'].fill = fundo_dia

        for row in range(4, 35):
            if row == 4:
                ws.cell(row=row, column=1).border = Border(
                    left=borda_ext.left,
                    right=borda_ext.right,
                    top=borda_ext.top,
                    bottom=borda_int.bottom
                )
            elif row == 34:
                ws.cell(row=row, column=1).border = Border(
                    left=borda_ext.left,
                    right=borda_ext.right,
                    top=borda_int.top,
                    bottom=borda_ext.bottom
                )
            else:
                ws.cell(row=row, column=1).border = Border(
                    left=borda_ext.left,
                    right=borda_ext.right,
                    top=borda_int.top,
                    bottom=borda_int.bottom
                )

        for row in range(2, 4):
            ws.cell(row=row, column=1).border = borda_dia

        colunas = [
            "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
            "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
        ]

        col_idx = 2
        total_qtd = [0] * 12
        total_valor = [Decimal('0.00')] * 12

        for i, mes in enumerate(colunas):
            fundo_mes = fundo_mes_claro if i % 2 == 0 else fundo_mes_escuro

            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
            ws.cell(row=2, column=col_idx, value=mes).font = fonte_negrito
            ws.cell(row=2, column=col_idx).alignment = alinhamento_central
            ws.cell(row=2, column=col_idx).fill = fundo_mes
            ws.cell(row=2, column=col_idx).border = borda_mes
            ws.cell(row=2, column=col_idx + 1).border = borda_mes

            ws.cell(row=3, column=col_idx, value="QTD").font = fonte_negrito
            ws.cell(row=3, column=col_idx).alignment = alinhamento_central
            ws.cell(row=3, column=col_idx).fill = fundo_mes
            ws.cell(row=3, column=col_idx).border = borda_qtd

            ws.cell(row=3, column=col_idx + 1, value="VALOR").font = fonte_negrito
            ws.cell(row=3, column=col_idx + 1).alignment = alinhamento_central
            ws.cell(row=3, column=col_idx + 1).fill = fundo_mes
            ws.cell(row=3, column=col_idx + 1).border = borda_valor

            col_idx += 2

        for dia in range(1, 32):
            ws.cell(row=dia + 3, column=1, value=dia).font = fonte_padrao
            ws.cell(row=dia + 3, column=1).fill = fundo_dia
            ws.cell(row=dia + 3, column=1).alignment = alinhamento_central

            col_idx = 2
            for mes in range(1, 13):
                censo_mes = censo_por_mes.get(mes, [])
                if dia <= len(censo_mes):
                    qtde = censo_mes[dia - 1]['qtde']
                    valor = censo_mes[dia - 1]['valor']
                else:
                    qtde = 0
                    valor = Decimal('0.00')

                ws.cell(row=dia + 3, column=col_idx, value=qtde).alignment = alinhamento_central
                valor_cell = ws.cell(row=dia + 3, column=col_idx + 1, value=valor)
                valor_cell.number_format = '#,##0.00'
                ws.cell(row=dia + 3, column=col_idx).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro
                ws.cell(row=dia + 3, column=col_idx + 1).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro

                total_qtd[mes - 1] += qtde
                total_valor[mes - 1] += Decimal(str(valor))
                col_idx += 2

        ws.cell(row=35, column=1, value="TOTAL").font = fonte_negrito
        ws.cell(row=35, column=1).alignment = alinhamento_central
        ws.cell(row=35, column=1).fill = fundo_dia

        col_idx = 2
        for mes in range(12):
            fundo_mes = fundo_mes_claro if mes % 2 == 0 else fundo_mes_escuro

            ws.cell(row=35, column=col_idx, value=total_qtd[mes]).font = fonte_negrito
            total_valor_cell = ws.cell(row=35, column=col_idx + 1, value=total_valor[mes])
            total_valor_cell.font = fonte_negrito
            total_valor_cell.number_format = '#,##0.00'
            ws.cell(row=35, column=col_idx).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx + 1).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx).fill = fundo_mes
            ws.cell(row=35, column=col_idx + 1).fill = fundo_mes

            col_idx += 2

        return wb

    @staticmethod
    def generate_censo_uh_ano_excel(ano, censo_htm_01, censo_htm_02):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Censo {ano}"

        fonte_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        fonte_padrao = Font(name="Calibri", size=10)
        fonte_negrito = Font(name="Calibri", size=10, bold=True)
        alinhamento_central = Alignment(horizontal="center", vertical="center")
        fundo_titulo = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        fundo_dia = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")
        fundo_mes_claro = PatternFill(start_color="F2F8EE", end_color="F2F8EE", fill_type="solid")
        fundo_mes_escuro = PatternFill(start_color="E2E0DA", end_color="E2E0DA", fill_type="solid")

        borda_titulo = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        borda_mes = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style=None)
        )
        borda_qtd = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )
        borda_uh = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style=None),
            bottom=Side(style='medium', color='000000')
        )

        ws.merge_cells('A1:Y1')
        titulo_cell = ws['A1']
        titulo_cell.value = "UNIDADES HABITACIONAIS"
        titulo_cell.font = fonte_titulo
        titulo_cell.alignment = alinhamento_central
        titulo_cell.fill = fundo_titulo
        titulo_cell.border = borda_titulo

        ws.merge_cells('A2:A3')
        ws['A2'] = "DIA"
        ws['A2'].font = fonte_negrito
        ws['A2'].alignment = alinhamento_central
        ws['A2'].fill = fundo_dia

        colunas = [
            "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
            "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
        ]

        col_idx = 2
        total_qtd = [0] * 12
        total_uh = [0] * 12

        for i, mes in enumerate(colunas):
            fundo_mes = fundo_mes_claro if i % 2 == 0 else fundo_mes_escuro

            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + 1)
            ws.cell(row=2, column=col_idx, value=mes).font = fonte_negrito
            ws.cell(row=2, column=col_idx).alignment = alinhamento_central
            ws.cell(row=2, column=col_idx).fill = fundo_mes
            ws.cell(row=2, column=col_idx).border = borda_mes
            ws.cell(row=2, column=col_idx + 1).border = borda_mes

            ws.cell(row=3, column=col_idx, value="TOTAL DE UH DO DIA").font = fonte_negrito
            ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=3, column=col_idx).fill = fundo_mes
            ws.cell(row=3, column=col_idx).border = borda_qtd

            ws.cell(row=3, column=col_idx + 1, value="UH OCUPADA NO DIA").font = fonte_negrito
            ws.cell(row=3, column=col_idx + 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=3, column=col_idx + 1).fill = fundo_mes
            ws.cell(row=3, column=col_idx + 1).border = borda_uh

            col_idx += 2

        for dia in range(1, 32):
            ws.cell(row=dia + 3, column=1, value=dia).font = fonte_padrao
            ws.cell(row=dia + 3, column=1).fill = fundo_dia
            ws.cell(row=dia + 3, column=1).alignment = alinhamento_central

            col_idx = 2
            for mes in range(1, 13):
                uh = 0
                if mes in censo_htm_01 and dia <= len(censo_htm_01[mes]):
                    uh += censo_htm_01[mes][dia - 1]['uh']
                if mes in censo_htm_02 and dia <= len(censo_htm_02[mes]):
                    uh += censo_htm_02[mes][dia - 1]['uh']

                ws.cell(row=dia + 3, column=col_idx, value="").alignment = alinhamento_central
                ws.cell(row=dia + 3, column=col_idx).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro
                ws.cell(row=dia + 3, column=col_idx).border = borda_qtd

                ws.cell(row=dia + 3, column=col_idx + 1, value=uh).alignment = alinhamento_central
                ws.cell(row=dia + 3, column=col_idx + 1).fill = fundo_mes_claro if mes % 2 == 1 else fundo_mes_escuro
                ws.cell(row=dia + 3, column=col_idx + 1).border = borda_uh

                total_uh[mes - 1] += uh
                col_idx += 2

        ws.cell(row=35, column=1, value="TOTAL").font = fonte_negrito
        ws.cell(row=35, column=1).alignment = alinhamento_central
        ws.cell(row=35, column=1).fill = fundo_dia

        col_idx = 2
        for mes in range(12):
            fundo_mes = fundo_mes_claro if mes % 2 == 0 else fundo_mes_escuro

            ws.cell(row=35, column=col_idx, value=total_qtd[mes]).font = fonte_negrito
            ws.cell(row=35, column=col_idx + 1, value=total_uh[mes]).font = fonte_negrito
            ws.cell(row=35, column=col_idx).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx + 1).alignment = alinhamento_central
            ws.cell(row=35, column=col_idx).fill = fundo_mes
            ws.cell(row=35, column=col_idx + 1).fill = fundo_mes

            col_idx += 2

        return wb
